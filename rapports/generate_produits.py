import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import date, timedelta

random.seed(99)

GREEN_DARK  = "1A6B3A"
GREEN_MED   = "2E7D32"
ORANGE      = "E65100"
WHITE       = "FFFFFF"
GREY_ROW    = "FAFAFA"
RED_LIGHT   = "FFEBEE"
YELLOW_L    = "FFFDE7"
GREEN_LIGHT = "E8F5E9"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="1A1A1A", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    t = Side(style="thin", color="DDDDDD")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def header_row(ws, row, cols):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = _fill(GREEN_DARK); c.font = _font(True, WHITE, 10)
        c.alignment = _align("center"); c.border = _border()

def data_row(ws, row, vals, alt=False, highlights=None):
    bg = GREY_ROW if alt else WHITE
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=val)
        cell_bg = highlights.get(i, bg) if highlights else bg
        c.fill = _fill(cell_bg); c.border = _border(); c.font = _font()
        if isinstance(val, (int, float)):
            c.alignment = _align("right")
            c.number_format = '#,##0' if val > 100 else '0'
        else:
            c.alignment = _align("left")

def banner(ws, title, sub, cols):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]; c.value = title
    c.fill = _fill(GREEN_DARK); c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = _align("center")
    ws.merge_cells(f"A2:{get_column_letter(cols)}2")
    c = ws["A2"]; c.value = sub
    c.fill = _fill(GREEN_MED); c.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    c.alignment = _align("center")

def widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Référentiels ──────────────────────────────────────────────
PRODUITS = [
    # (code, nom, catégorie, unité, prix_achat, prix_vente, stock_min, fournisseur_principal)
    ("PRD-001", "Riz local 50 kg",          "Alimentation",   "Sac",     22000, 25000,  20, "SANIA CI"),
    ("PRD-002", "Riz importé 25 kg",         "Alimentation",   "Sac",     16000, 18500,  15, "SANIA CI"),
    ("PRD-003", "Huile de palme 5 L",        "Alimentation",   "Bidon",    3800,  4500,  30, "SIFCA"),
    ("PRD-004", "Sucre en poudre 50 kg",     "Alimentation",   "Sac",     24500, 28000,  10, "SUCRIVOIRE"),
    ("PRD-005", "Farine de blé 25 kg",       "Alimentation",   "Sac",     10500, 12000,  20, "Grand Moulin CI"),
    ("PRD-006", "Lait en poudre 2,5 kg",     "Alimentation",   "Boîte",    7200,  8500,  25, "NESTLÉ CI"),
    ("PRD-007", "Tomate concentrée x24",     "Alimentation",   "Carton",  13000, 15000,  15, "CFCI"),
    ("PRD-008", "Sardines conserve x24",     "Alimentation",   "Carton",  19000, 22000,  10, "CFCI"),
    ("PRD-009", "Café soluble 500 g",        "Boissons",       "Boîte",    2800,  3500,  20, "NESTLÉ CI"),
    ("PRD-010", "Eau minérale x12 1,5 L",   "Boissons",       "Carton",   4200,  5500,  40, "NESTLÉ CI"),
    ("PRD-011", "Jus de fruit 1 L x12",     "Boissons",       "Carton",   8000,  9500,  15, "Traoré Dist."),
    ("PRD-012", "Savon de ménage x12",       "Hygiène",        "Carton",   5000,  6000,  20, "UNILEVER CI"),
    ("PRD-013", "Détergent 5 kg",            "Hygiène",        "Sac",      7500,  9000,  15, "UNILEVER CI"),
    ("PRD-014", "Lessive en poudre 5 kg",   "Hygiène",        "Sac",      6200,  7500,  15, "UNILEVER CI"),
    ("PRD-015", "Gel douche 250 ml x12",    "Hygiène",        "Carton",   9500, 11000,  10, "UNILEVER CI"),
    ("PRD-016", "Huile moteur 5 L",          "Quincaillerie",  "Bidon",   15000, 18000,   8, "PETROCI"),
    ("PRD-017", "Pile alcaline AA x12",     "Quincaillerie",  "Carton",   2200,  3000,  20, "ElecPlus CI"),
    ("PRD-018", "Ampoule LED 9W x10",       "Quincaillerie",  "Boîte",    5500,  7000,  10, "ElecPlus CI"),
    ("PRD-019", "Sel iodé 1 kg x20",        "Alimentation",   "Carton",   4000,  5000,  25, "Diallo Commerce"),
    ("PRD-020", "Cube Maggi x120",          "Alimentation",   "Carton",   6500,  8000,  20, "NESTLÉ CI"),
]

DEPOTS = ["Dépôt Principal", "Dépôt Yopougon", "Dépôt Cocody"]

MOTIFS_ENTREE  = ["Approvisionnement", "Retour client", "Transfert entrant"]
MOTIFS_SORTIE  = ["Vente", "Vente", "Vente", "Transfert sortant"]
MOTIFS_PERTE   = ["Produit périmé", "Casse", "Vol constaté"]
MOTIFS_INVENT  = ["Inventaire périodique", "Inventaire annuel"]

# Stocks par dépôt (simulés, cohérents avec les ventes)
stock_data = {}
for p in PRODUITS:
    code = p[0]
    stock_data[code] = {
        "Dépôt Principal": random.randint(p[6], p[6] * 8),
        "Dépôt Yopougon":  random.randint(p[6] // 2, p[6] * 5),
        "Dépôt Cocody":    random.randint(p[6] // 2, p[6] * 4),
    }


# ════════════════════════════════════════════════════════════════
#  FEUILLE 1 — CATALOGUE PRODUITS
# ════════════════════════════════════════════════════════════════
def sheet_catalogue(wb):
    ws = wb.create_sheet("Catalogue Produits")
    ws.freeze_panes = "A5"
    cols = ["Code", "Désignation", "Catégorie", "Unité",
            "Prix achat (FCFA)", "Prix vente (FCFA)", "Marge (FCFA)", "Marge (%)",
            "Stock mini", "Fournisseur principal", "Statut"]
    banner(ws, "CATALOGUE DES PRODUITS",
           f"AGSV – Référentiel produits  |  Mis à jour le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    widths(ws, [10, 28, 16, 10, 18, 18, 14, 10, 10, 22, 10])

    for idx, p in enumerate(PRODUITS):
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        marge    = p_vte - p_ach
        marge_pct = round(marge / p_ach * 100, 1)
        stock_total = sum(stock_data[code].values())
        statut = "ACTIF" if stock_total > s_min else "RUPTURE"
        hl = {11: GREEN_LIGHT if statut == "ACTIF" else RED_LIGHT}
        data_row(ws, 5 + idx,
                 [code, nom, cat, unit, p_ach, p_vte, marge, f"{marge_pct} %",
                  s_min, fourn, statut],
                 alt=(idx % 2 == 1), highlights=hl)
        ws.row_dimensions[5 + idx].height = 20

    # Totaux
    tr = 5 + len(PRODUITS)
    ws.merge_cells(f"A{tr}:D{tr}")
    ws[f"A{tr}"].value = f"TOTAL  –  {len(PRODUITS)} références actives"
    ws[f"A{tr}"].font  = _font(True, WHITE); ws[f"A{tr}"].fill = _fill(GREEN_DARK)
    ws[f"A{tr}"].alignment = _align("center")
    for col in range(5, 12):
        ws.cell(row=tr, column=col).fill = _fill(GREEN_DARK)


# ════════════════════════════════════════════════════════════════
#  FEUILLE 2 — STOCK PAR DÉPÔT
# ════════════════════════════════════════════════════════════════
def sheet_stock(wb):
    ws = wb.create_sheet("Stock par Dépôt")
    ws.freeze_panes = "A5"
    cols = ["Code", "Désignation", "Catégorie",
            "Dépôt Principal", "Dépôt Yopougon", "Dépôt Cocody",
            "Stock total", "Valeur stock (FCFA)", "Stock mini", "Alerte"]
    banner(ws, "ÉTAT DES STOCKS PAR DÉPÔT",
           f"AGSV – Situation au {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    widths(ws, [10, 28, 16, 16, 16, 14, 12, 22, 10, 12])

    val_total = 0
    for idx, p in enumerate(PRODUITS):
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        dp = stock_data[code]["Dépôt Principal"]
        dy = stock_data[code]["Dépôt Yopougon"]
        dc = stock_data[code]["Dépôt Cocody"]
        total = dp + dy + dc
        valeur = total * p_vte
        val_total += valeur
        alerte = "⚠ STOCK BAS" if total <= s_min * 1.5 else ("✓ OK" if total > s_min * 3 else "→ MOYEN")
        hl = {10: RED_LIGHT if "⚠" in alerte else (GREEN_LIGHT if "✓" in alerte else YELLOW_L)}
        data_row(ws, 5 + idx, [code, nom, cat, dp, dy, dc, total, valeur, s_min, alerte],
                 alt=(idx % 2 == 1), highlights=hl)
        ws.row_dimensions[5 + idx].height = 20

    tr = 5 + len(PRODUITS)
    ws.merge_cells(f"A{tr}:G{tr}")
    ws[f"A{tr}"].value = "VALEUR TOTALE DU STOCK"
    ws[f"A{tr}"].font  = _font(True, WHITE); ws[f"A{tr}"].fill = _fill(GREEN_DARK)
    ws[f"A{tr}"].alignment = _align("center")
    ws.cell(row=tr, column=8).value = val_total
    ws.cell(row=tr, column=8).font  = _font(True, WHITE)
    ws.cell(row=tr, column=8).fill  = _fill(GREEN_DARK)
    ws.cell(row=tr, column=8).alignment = _align("right")
    ws.cell(row=tr, column=8).number_format = '#,##0'
    for col in [9, 10]:
        ws.cell(row=tr, column=col).fill = _fill(GREEN_DARK)


# ════════════════════════════════════════════════════════════════
#  FEUILLE 3 — MOUVEMENTS DE STOCK
# ════════════════════════════════════════════════════════════════
def sheet_mouvements(wb):
    ws = wb.create_sheet("Mouvements de Stock")
    ws.freeze_panes = "A5"
    cols = ["N° Mouvement", "Date", "Produit", "Catégorie",
            "Type", "Quantité", "Motif", "Dépôt", "Référence"]
    banner(ws, "HISTORIQUE DES MOUVEMENTS DE STOCK – 2025/2026",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols))
    header_row(ws, 4, cols)
    widths(ws, [14, 12, 28, 16, 12, 10, 28, 20, 16])

    mvts = []
    for year in [2025, 2026]:
        n = 300 if year == 2025 else 120
        for i in range(n):
            d   = _rand_date(year)
            p   = random.choice(PRODUITS)
            typ = random.choices(["ENTRÉE","SORTIE","PERTE","INVENTAIRE"],
                                 weights=[25, 55, 10, 10])[0]
            qty = random.randint(1, 50)
            if typ == "ENTRÉE":   motif = random.choice(MOTIFS_ENTREE)
            elif typ == "SORTIE": motif = random.choice(MOTIFS_SORTIE)
            elif typ == "PERTE":  motif = random.choice(MOTIFS_PERTE)
            else:                 motif = random.choice(MOTIFS_INVENT)
            depot = random.choice(DEPOTS)
            ref   = f"MVT-{year}-{i+1:04d}"
            mvts.append((ref, d, p[1], p[2], typ, qty, motif, depot, ref.replace("MVT","REF")))

    mvts.sort(key=lambda x: x[1])

    TYPE_COLORS = {"ENTRÉE": GREEN_LIGHT, "SORTIE": WHITE, "PERTE": RED_LIGHT, "INVENTAIRE": YELLOW_L}
    for idx, vals in enumerate(mvts):
        typ_color = TYPE_COLORS.get(vals[4], WHITE)
        hl = {5: typ_color}
        data_row(ws, 5 + idx, vals, alt=(idx % 2 == 1), highlights=hl)
        ws.row_dimensions[5 + idx].height = 18


# ════════════════════════════════════════════════════════════════
#  FEUILLE 4 — PERFORMANCE PRODUITS
# ════════════════════════════════════════════════════════════════
def sheet_performance(wb):
    ws = wb.create_sheet("Performance Produits")
    ws.freeze_panes = "A5"
    cols = ["Code", "Désignation", "Catégorie",
            "Qté vendue 2025", "CA 2025 (FCFA)", "Qté vendue 2026 (jan–mai)", "CA 2026 (FCFA)",
            "Évolution CA", "Rang 2025"]
    banner(ws, "PERFORMANCE DES VENTES PAR PRODUIT",
           f"AGSV – Analyse 2025 vs 2026 (jan–mai)  |  Généré le {date.today():%d/%m/%Y}",
           len(cols))
    header_row(ws, 4, cols)
    widths(ws, [10, 28, 16, 18, 20, 24, 20, 14, 10])

    perfs = []
    for p in PRODUITS:
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        qty_2025 = random.randint(50, 600)
        ca_2025  = qty_2025 * p_vte
        qty_2026 = random.randint(20, int(qty_2025 * 0.55))
        ca_2026  = qty_2026 * p_vte
        annualise = ca_2026 * 12 / 5
        evol = round((annualise - ca_2025) / ca_2025 * 100, 1)
        perfs.append((code, nom, cat, qty_2025, ca_2025, qty_2026, ca_2026, evol))

    perfs.sort(key=lambda x: -x[4])

    for idx, (code, nom, cat, q25, ca25, q26, ca26, evol) in enumerate(perfs):
        rang = idx + 1
        evol_str = f"▲ +{evol} %" if evol >= 0 else f"▼ {evol} %"
        hl = {8: GREEN_LIGHT if evol >= 0 else RED_LIGHT}
        data_row(ws, 5 + idx, [code, nom, cat, q25, ca25, q26, ca26, evol_str, rang],
                 alt=(idx % 2 == 1), highlights=hl)
        ws.row_dimensions[5 + idx].height = 20

    tr = 5 + len(perfs)
    ws.merge_cells(f"A{tr}:C{tr}")
    ws[f"A{tr}"].value = "TOTAUX"
    ws[f"A{tr}"].font  = _font(True, WHITE); ws[f"A{tr}"].fill = _fill(GREEN_DARK)
    ws[f"A{tr}"].alignment = _align("center")
    for col_idx, total in enumerate([sum(x[3] for x in perfs),
                                      sum(x[4] for x in perfs),
                                      sum(x[5] for x in perfs),
                                      sum(x[6] for x in perfs)], start=4):
        c = ws.cell(row=tr, column=col_idx, value=total)
        c.font = _font(True, WHITE); c.fill = _fill(GREEN_DARK)
        c.alignment = _align("right"); c.number_format = '#,##0'
    for col in [8, 9]:
        ws.cell(row=tr, column=col).fill = _fill(GREEN_DARK)


# ════════════════════════════════════════════════════════════════
#  FEUILLE 5 — RÉSUMÉ PAR CATÉGORIE
# ════════════════════════════════════════════════════════════════
def sheet_categories(wb):
    ws = wb.create_sheet("Résumé par Catégorie")
    ws.freeze_panes = "A5"
    cols = ["Catégorie", "Nb références", "Valeur stock (FCFA)",
            "CA 2025 (FCFA)", "CA 2026 jan–mai (FCFA)", "Marge moyenne (%)"]
    banner(ws, "ANALYSE PAR CATÉGORIE DE PRODUITS",
           f"AGSV – Synthèse  |  Généré le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    widths(ws, [20, 16, 24, 24, 26, 18])

    cats = {}
    for p in PRODUITS:
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        stock_t = sum(stock_data[code].values())
        marge   = round((p_vte - p_ach) / p_ach * 100, 1)
        if cat not in cats:
            cats[cat] = {"nb": 0, "val_stock": 0, "ca25": 0, "ca26": 0, "marges": []}
        cats[cat]["nb"]        += 1
        cats[cat]["val_stock"] += stock_t * p_vte
        cats[cat]["ca25"]      += random.randint(200000, 2000000)
        cats[cat]["ca26"]      += random.randint(80000, 900000)
        cats[cat]["marges"].append(marge)

    for idx, (cat, d) in enumerate(sorted(cats.items())):
        moy_marge = round(sum(d["marges"]) / len(d["marges"]), 1)
        data_row(ws, 5 + idx,
                 [cat, d["nb"], d["val_stock"], d["ca25"], d["ca26"], f"{moy_marge} %"],
                 alt=(idx % 2 == 1))
        ws.row_dimensions[5 + idx].height = 22

    tr = 5 + len(cats)
    ws.merge_cells(f"A{tr}:B{tr}")
    ws[f"A{tr}"].value = "TOTAL"
    ws[f"A{tr}"].font  = _font(True, WHITE); ws[f"A{tr}"].fill = _fill(GREEN_DARK)
    ws[f"A{tr}"].alignment = _align("center")
    for col_idx, total in enumerate([
        sum(d["val_stock"] for d in cats.values()),
        sum(d["ca25"]      for d in cats.values()),
        sum(d["ca26"]      for d in cats.values()),
    ], start=3):
        c = ws.cell(row=tr, column=col_idx, value=total)
        c.font = _font(True, WHITE); c.fill = _fill(GREEN_DARK)
        c.alignment = _align("right"); c.number_format = '#,##0'
    ws.cell(row=tr, column=6).fill = _fill(GREEN_DARK)


def _rand_date(year):
    month   = random.randint(1, 12 if year == 2025 else 5)
    day_max = 28 if month == 2 else (30 if month in [4,6,9,11] else 31)
    return date(year, month, random.randint(1, day_max))


# ── Génération ────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

sheet_catalogue(wb)
sheet_stock(wb)
sheet_mouvements(wb)
sheet_performance(wb)
sheet_categories(wb)

out = "/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Produits.xlsx"
wb.save(out)
print(f"✓ Rapport_Produits.xlsx  ({5} feuilles)")