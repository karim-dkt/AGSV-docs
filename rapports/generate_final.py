"""
Génération des rapports Excel COOPRANORD
  - Ventes Korhogo 2025-2026
  - Ventes Bamako  2025-2026
  - Commandes Korhogo 2025-2026
  - Commandes Bamako  2025-2026
  - Produits (catalogue partagé entreprise)
"""
import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Palette ────────────────────────────────────────────────────────────────────
VERT_FONCE  = "1A6B3A"
VERT_CLAIR  = "D6EAD7"
ORANGE      = "E65100"
ORANGE_PALE = "FBE9E7"
GRIS_LIGNE  = "F5F5F5"
BLANC       = "FFFFFF"

def side(color="CCCCCC"):
    return Side(style="thin", color=color)

BORDER_LIGHT  = Border(left=side(), right=side(), top=side(), bottom=side())
BORDER_HEADER = Border(
    left=side(VERT_FONCE), right=side(VERT_FONCE),
    top=side(VERT_FONCE),  bottom=side(VERT_FONCE)
)

def hfont(bold=True, color=BLANC, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cfont(bold=False, color="222222", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ── Catalogue produits COOPRANORD ──────────────────────────────────────────────
# (ref, nom, catégorie, unité, prix_achat, prix_vente, stock_korhogo, stock_bamako, seuil_alerte)
PRODUITS = [
    # Arbres fruitiers
    ("AF-001", "Prunier",                    "Arbres fruitiers",               "unité",  3800,  5000,  420, 310, 100),
    ("AF-002", "Pommier",                    "Arbres fruitiers",               "unité",  5500,  7500,  380, 260,  80),
    ("AF-003", "Poirier",                    "Arbres fruitiers",               "unité",  4800,  6500,  290, 195,  60),
    ("AF-004", "Pêcher",                     "Arbres fruitiers",               "unité",  4400,  6000,  350, 230,  80),
    ("AF-005", "Figuier turc",               "Arbres fruitiers",               "unité",  3200,  4500,  510, 370, 100),
    ("AF-006", "Kaki",                       "Arbres fruitiers",               "unité",  4000,  5500,  270, 180,  60),
    ("AF-007", "Nèflier",                    "Arbres fruitiers",               "unité",  2800,  4000,  460, 300,  80),
    ("AF-008", "Figuier commun",             "Arbres fruitiers",               "unité",  2500,  3500,  600, 420, 120),
    ("AF-009", "Fruit du dragon (Pitaya)",   "Arbres fruitiers",               "unité",  5800,  8000,  200, 140,  50),
    ("AF-010", "Anone (pomme à la crème)",   "Arbres fruitiers",               "unité",  3200,  4500,  330, 220,  80),
    ("AF-011", "Goyavier",                   "Arbres fruitiers",               "unité",  2200,  3000,  700, 510, 150),
    ("AF-012", "Vigne (Raisins)",            "Arbres fruitiers",               "unité",  3500,  5000,  280, 190,  60),
    ("AF-013", "Caroubier",                  "Arbres fruitiers",               "unité",  4300,  6000,  160,  90,  40),
    ("AF-014", "Avocatier",                  "Arbres fruitiers",               "unité",  2800,  4000,  850, 640, 200),
    ("AF-015", "Limequat",                   "Arbres fruitiers",               "unité",  5000,  7000,  120,  80,  30),
    ("AF-016", "Kumquat",                    "Arbres fruitiers",               "unité",  5400,  7500,  110,  75,  30),
    ("AF-017", "Citron caviar",              "Arbres fruitiers",               "unité",  8800, 12000,   80,  50,  20),
    ("AF-018", "Grenadier",                  "Arbres fruitiers",               "unité",  4000,  5500,  410, 280,  80),
    ("AF-019", "Cerisier",                   "Arbres fruitiers",               "unité",  6500,  9000,   90,  55,  20),
    ("AF-020", "Abricotier",                 "Arbres fruitiers",               "unité",  6000,  8500,  130,  85,  30),
    ("AF-021", "Cognassier",                 "Arbres fruitiers",               "unité",  4700,  6500,  175, 110,  40),
    ("AF-022", "Nectarinier",                "Arbres fruitiers",               "unité",  5000,  7000,  145,  95,  30),
    ("AF-023", "Pamplemoussier",             "Arbres fruitiers",               "unité",  3600,  5000,  320, 215,  60),
    # Plantes décoratives & aromatiques
    ("PD-001", "Feuille de bananier (décoratif)", "Plantes décoratives & aromatiques", "unité",  900, 1500, 1200, 980, 300),
    ("PD-002", "Framboisier",                "Plantes décoratives & aromatiques", "unité",  2800,  4000,  390, 260,  80),
    ("PD-003", "Moringa",                    "Plantes décoratives & aromatiques", "unité",  1700,  2500,  870, 620, 200),
    ("PD-004", "Cyprès (décoratif)",         "Plantes décoratives & aromatiques", "unité",  2500,  3500,  440, 310, 100),
    ("PD-005", "Lavande",                    "Plantes décoratives & aromatiques", "unité",  1400,  2000,  760, 540, 150),
    ("PD-006", "Thym",                       "Plantes décoratives & aromatiques", "unité",   900,  1500,  920, 700, 200),
    ("PD-007", "Armoise",                    "Plantes décoratives & aromatiques", "unité",  1200,  1800,  680, 480, 150),
    # Céréales & oléagineux (CDC)
    ("CO-001", "Beurre de karité",           "Céréales & oléagineux",          "kg",  1800,  2500, 2800, 3200, 500),
    ("CO-002", "Grain de karité",            "Céréales & oléagineux",          "kg",   280,   400, 8500, 9200, 2000),
    ("CO-003", "Mil",                        "Céréales & oléagineux",          "kg",   240,   350, 12000, 15000, 3000),
    ("CO-004", "Sorgho",                     "Céréales & oléagineux",          "kg",   220,   320, 10500, 13000, 2500),
    ("CO-005", "Maïs",                       "Céréales & oléagineux",          "kg",   195,   280, 14000, 18000, 4000),
]

# ── Clients (personnes physiques uniquement) ───────────────────────────────────
CLIENTS_KORHOGO = [
    "Coulibaly Mamadou", "Koné Fatimata",    "Traoré Ibrahim",
    "Bamba Aminata",     "Kouyaté Seydou",   "Sangaré Mariam",
    "Ouattara Lassina",  "Diabaté Fatoumata","Soro Boubacar",
    "Koulibaly Nadia",   "Bah Adama",        "Fofana Kadiatou",
]

CLIENTS_BAMAKO = [
    "Diallo Oumar",      "Keita Aissatou",   "Coulibaly Moussa",
    "Traoré Fatoumata",  "Maiga Issouf",     "Cissé Safiatou",
    "Touré Boubacar",    "Diakité Aminatou", "Sidibé Moussa",
    "Bathily Awa",       "Dembélé Salif",    "Konaré Rokia",
]

VENDEURS_KORHOGO = ["Koné Drissa", "Bamba Mariam"]
VENDEURS_BAMAKO  = ["Diallo Seydou", "Traoré Aminata"]

MODES_PAIEMENT   = ["Espèces", "Mobile Money", "Lettre d'échange"]
STATUTS_VENTE    = ["Payée", "Crédit", "Partiellement payée"]
STATUTS_COMMANDE = ["Livrée", "En attente", "En cours", "Annulée"]

# ── Helpers ────────────────────────────────────────────────────────────────────
def random_date(year):
    start = date(year, 1, 1)
    end   = date(year, 12, 31)
    return start + timedelta(days=random.randint(0, (end - start).days))

def fmt_date(d):
    return d.strftime("%d/%m/%Y")

def set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def apply_header(ws, row_idx, values, widths):
    for ci, (val, w) in enumerate(zip(values, widths), 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.font      = hfont()
        cell.fill      = fill(VERT_FONCE)
        cell.alignment = CENTER
        cell.border    = BORDER_HEADER
        set_col_width(ws, ci, w)

def apply_data_row(ws, row_idx, values, number_cols=None, currency_cols=None):
    bg = GRIS_LIGNE if row_idx % 2 == 0 else BLANC
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.font   = cfont()
        cell.fill   = fill(bg)
        cell.border = BORDER_LIGHT
        if number_cols and ci in number_cols:
            cell.alignment    = RIGHT
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        elif currency_cols and ci in currency_cols:
            cell.alignment    = RIGHT
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0" FCFA"'
        else:
            cell.alignment = LEFT

def title_row(ws, text, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font      = Font(name="Calibri", bold=True, color=BLANC, size=13)
    cell.fill      = fill(VERT_FONCE)
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

def subtitle_row(ws, text, n_cols, row=2):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = Font(name="Calibri", bold=True, color=ORANGE, size=11)
    cell.fill      = fill(ORANGE_PALE)
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 20

def total_row(ws, row_idx, n_cols, total_label, total_val, total_col):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_col - 1)
    lbl = ws.cell(row=row_idx, column=1, value=total_label)
    lbl.font      = Font(name="Calibri", bold=True, color=VERT_FONCE, size=10)
    lbl.fill      = fill(VERT_CLAIR)
    lbl.alignment = RIGHT

    val = ws.cell(row=row_idx, column=total_col, value=total_val)
    val.font          = Font(name="Calibri", bold=True, color=VERT_FONCE, size=10)
    val.fill          = fill(VERT_CLAIR)
    val.alignment     = RIGHT
    val.number_format = '#,##0" FCFA"'

    for c in range(total_col + 1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill   = fill(VERT_CLAIR)
        cell.border = BORDER_LIGHT

    ws.row_dimensions[row_idx].height = 18

# ── Ventes ─────────────────────────────────────────────────────────────────────
VENTE_COLS   = ["N° Vente", "Date", "Client", "Produit", "Catégorie", "Unité",
                "Quantité", "Prix unitaire (FCFA)", "Total (FCFA)",
                "Mode de paiement", "Statut", "Vendeur"]
VENTE_WIDTHS = [14, 14, 24, 30, 30, 10, 12, 22, 22, 22, 22, 22]

PRODUITS_PLANTES   = [p for p in PRODUITS if p[3] == "unité"]
PRODUITS_CEREALES  = [p for p in PRODUITS if p[3] == "kg"]

def generer_lignes_ventes(clients, vendeurs, n, year):
    # 2025 = exercice clôturé : seuls Payée ou Crédit (dette historique)
    # 2026 = exercice en cours : tous les statuts possibles
    statuts_dispo = ["Payée", "Crédit"] if year < 2026 else STATUTS_VENTE
    lignes = []
    for i in range(1, n + 1):
        d      = random_date(year)
        client = random.choice(clients)
        if random.random() < 0.55:
            prod = random.choice(PRODUITS_PLANTES)
            qty  = random.randint(50, 500)
        else:
            prod = random.choice(PRODUITS_CEREALES)
            qty  = random.randint(500, 8000)
        prix   = prod[5]
        total  = qty * prix
        mode   = random.choice(MODES_PAIEMENT)
        statut = random.choice(statuts_dispo)
        vendeur = random.choice(vendeurs)
        lignes.append((
            f"VTE-{year}-{i:04d}", fmt_date(d), client,
            prod[1], prod[2], prod[3],
            qty, prix, total, mode, statut, vendeur
        ))
    lignes.sort(key=lambda x: x[1])
    return lignes

def generer_ventes(depot_key):
    depot_info = {
        "korhogo": ("Korhogo — Côte d'Ivoire", CLIENTS_KORHOGO, VENDEURS_KORHOGO, "Korhogo"),
        "bamako":  ("Bamako — Mali",            CLIENTS_BAMAKO,  VENDEURS_BAMAKO,  "Bamako"),
    }
    label, clients, vendeurs, fname = depot_info[depot_key]

    wb = Workbook()
    wb.remove(wb.active)

    for year in [2025, 2026]:
        n      = random.randint(28, 38)
        lignes = generer_lignes_ventes(clients, vendeurs, n, year)
        total_ca = sum(l[8] for l in lignes)

        ws = wb.create_sheet(title=str(year))
        ws.freeze_panes = "A5"

        n_cols = len(VENTE_COLS)
        title_row(ws, f"COOPRANORD — Rapport des Ventes — Dépôt de {label}", n_cols)
        subtitle_row(ws, f"Exercice {year}  |  {len(lignes)} ventes  |  CA total : {total_ca:,.0f} FCFA", n_cols)
        ws.row_dimensions[3].height = 6
        apply_header(ws, 4, VENTE_COLS, VENTE_WIDTHS)

        for ri, ligne in enumerate(lignes, 5):
            apply_data_row(ws, ri, ligne,
                           number_cols={7}, currency_cols={8, 9})

        last = 4 + len(lignes)
        total_row(ws, last + 1, n_cols,
                  f"CHIFFRE D'AFFAIRES TOTAL {year}", total_ca, 9)

    wb.save(f"Rapport_Ventes_{fname}_2025-2026.xlsx")
    print(f"  ✓ Rapport_Ventes_{fname}_2025-2026.xlsx")

# ── Commandes ──────────────────────────────────────────────────────────────────
COMMANDE_COLS   = ["N° Commande", "Date commande", "Client", "Produit", "Catégorie", "Unité",
                   "Quantité", "Prix unitaire (FCFA)", "Total (FCFA)",
                   "Date de livraison", "Statut", "Vendeur"]
COMMANDE_WIDTHS = [16, 16, 24, 30, 30, 10, 12, 22, 22, 18, 18, 22]

def generer_lignes_commandes(clients, vendeurs, n, year):
    # 2025 = exercice clôturé : seuls Livrée ou Annulée
    # 2026 = exercice en cours : tous les statuts possibles
    statuts_dispo = ["Livrée", "Annulée"] if year < 2026 else STATUTS_COMMANDE
    lignes = []
    for i in range(1, n + 1):
        d_cmd  = random_date(year)
        d_livr = d_cmd + timedelta(days=random.randint(3, 21))
        client = random.choice(clients)
        if random.random() < 0.55:
            prod = random.choice(PRODUITS_PLANTES)
            qty  = random.randint(100, 600)
        else:
            prod = random.choice(PRODUITS_CEREALES)
            qty  = random.randint(1000, 10000)
        prix   = prod[5]
        total  = qty * prix
        statut = random.choice(statuts_dispo)
        vendeur = random.choice(vendeurs)
        lignes.append((
            f"CMD-{year}-{i:04d}", fmt_date(d_cmd), client,
            prod[1], prod[2], prod[3],
            qty, prix, total,
            fmt_date(d_livr) if statut != "Annulée" else "—",
            statut, vendeur
        ))
    lignes.sort(key=lambda x: x[1])
    return lignes

def generer_commandes(depot_key):
    depot_info = {
        "korhogo": ("Korhogo — Côte d'Ivoire", CLIENTS_KORHOGO, VENDEURS_KORHOGO, "Korhogo"),
        "bamako":  ("Bamako — Mali",            CLIENTS_BAMAKO,  VENDEURS_BAMAKO,  "Bamako"),
    }
    label, clients, vendeurs, fname = depot_info[depot_key]

    wb = Workbook()
    wb.remove(wb.active)

    for year in [2025, 2026]:
        n      = random.randint(22, 32)
        lignes = generer_lignes_commandes(clients, vendeurs, n, year)
        total_val = sum(l[8] for l in lignes if l[10] != "Annulée")

        ws = wb.create_sheet(title=str(year))
        ws.freeze_panes = "A5"

        n_cols = len(COMMANDE_COLS)
        title_row(ws, f"COOPRANORD — Rapport des Commandes — Dépôt de {label}", n_cols)
        subtitle_row(ws, f"Exercice {year}  |  {len(lignes)} commandes  |  Valeur livrée : {total_val:,.0f} FCFA", n_cols)
        ws.row_dimensions[3].height = 6
        apply_header(ws, 4, COMMANDE_COLS, COMMANDE_WIDTHS)

        for ri, ligne in enumerate(lignes, 5):
            apply_data_row(ws, ri, ligne,
                           number_cols={7}, currency_cols={8, 9})

        last = 4 + len(lignes)
        total_row(ws, last + 1, n_cols,
                  f"VALEUR TOTALE LIVRÉE {year}", total_val, 9)

    wb.save(f"Rapport_Commandes_{fname}_2025-2026.xlsx")
    print(f"  ✓ Rapport_Commandes_{fname}_2025-2026.xlsx")

# ── Produits (catalogue entreprise, partagé) ───────────────────────────────────
PROD_COLS   = ["Référence", "Nom du produit", "Catégorie", "Unité",
               "Prix d'achat (FCFA)", "Prix de vente (FCFA)",
               "Stock Korhogo", "Stock Bamako", "Stock Total",
               "Seuil d'alerte", "Statut"]
PROD_WIDTHS = [12, 34, 34, 10, 22, 22, 16, 14, 14, 16, 14]

def statut_stock(stock, seuil):
    if stock == 0:      return "Rupture"
    if stock <= seuil:  return "Critique"
    if stock <= seuil * 1.5: return "Faible"
    return "Normal"

def generer_produits():
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalogue produits"
    ws.freeze_panes = "A5"

    n_cols = len(PROD_COLS)
    title_row(ws, "COOPRANORD — Catalogue des Produits — Exercice 2025-2026", n_cols)
    subtitle_row(ws, f"{len(PRODUITS)} références  |  Catalogue commun — Dépôts : Korhogo & Bamako", n_cols)
    ws.row_dimensions[3].height = 6
    apply_header(ws, 4, PROD_COLS, PROD_WIDTHS)

    # Grouper par catégorie
    cats_ordered = []
    for p in PRODUITS:
        if p[2] not in cats_ordered:
            cats_ordered.append(p[2])

    row = 5
    for cat in cats_ordered:
        prods_cat = [p for p in PRODUITS if p[2] == cat]

        # Ligne d'entête catégorie
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cell = ws.cell(row=row, column=1, value=f"  {cat.upper()}")
        cell.font      = Font(name="Calibri", bold=True, color=BLANC, size=10)
        cell.fill      = fill(ORANGE)
        cell.alignment = LEFT
        ws.row_dimensions[row].height = 16
        row += 1

        for p in prods_cat:
            ref, nom, categorie, unite, prix_achat, prix_vente, stk_kor, stk_bam, seuil = p
            stk_total = stk_kor + stk_bam
            seuil_total = seuil * 2
            statut = statut_stock(stk_total, seuil_total)
            ligne = (ref, nom, categorie, unite,
                     prix_achat, prix_vente,
                     stk_kor, stk_bam, stk_total,
                     seuil_total, statut)
            apply_data_row(ws, row, ligne,
                           number_cols={7, 8, 9, 10},
                           currency_cols={5, 6})

            # Colorer le statut
            sc = ws.cell(row=row, column=11)
            colors = {"Rupture": "C62828", "Critique": ORANGE, "Faible": "F9A825", "Normal": VERT_FONCE}
            sc.font      = Font(name="Calibri", bold=True, color=colors[statut], size=10)
            sc.alignment = CENTER
            row += 1

    # Ligne totaux
    total_kor = sum(p[6] for p in PRODUITS)
    total_bam = sum(p[7] for p in PRODUITS)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    lbl = ws.cell(row=row, column=1, value="STOCK TOTAL ENTREPRISE")
    lbl.font      = Font(name="Calibri", bold=True, color=VERT_FONCE, size=10)
    lbl.fill      = fill(VERT_CLAIR)
    lbl.alignment = RIGHT

    for ci, val in [(7, total_kor), (8, total_bam), (9, total_kor + total_bam)]:
        c = ws.cell(row=row, column=ci, value=val)
        c.font          = Font(name="Calibri", bold=True, color=VERT_FONCE, size=10)
        c.fill          = fill(VERT_CLAIR)
        c.alignment     = RIGHT
        c.number_format = '#,##0'
        c.border        = BORDER_LIGHT

    for ci in [10, 11]:
        c = ws.cell(row=row, column=ci)
        c.fill   = fill(VERT_CLAIR)
        c.border = BORDER_LIGHT

    ws.row_dimensions[row].height = 18

    wb.save("Rapport_Produits.xlsx")
    print("  ✓ Rapport_Produits.xlsx")

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("── Ventes (par dépôt) ──")
    generer_ventes("korhogo")
    generer_ventes("bamako")

    print("\n── Commandes (par dépôt) ──")
    generer_commandes("korhogo")
    generer_commandes("bamako")

    print("\n── Catalogue produits ──")
    generer_produits()

    print("\nTerminé.")
