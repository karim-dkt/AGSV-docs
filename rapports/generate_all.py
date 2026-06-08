import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import date, timedelta

random.seed(42)

# ── Styles ────────────────────────────────────────────────────
GREEN_DARK  = "1A6B3A"
GREEN_MED   = "2E7D32"
GREEN_LIGHT = "E8F5E9"
ORANGE      = "E65100"
WHITE       = "FFFFFF"
GREY_ROW    = "FAFAFA"
RED_LIGHT   = "FFEBEE"
YELLOW_L    = "FFFDE7"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="1A1A1A", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    t = Side(style="thin", color="DDDDDD")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def header_row(ws, row, cols):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = _fill(GREEN_DARK); c.font = _font(True, WHITE, 10)
        c.alignment = _align("center"); c.border = _border()

def data_row(ws, row, vals, alt=False, highlights=None):
    bg = GREY_ROW if alt else WHITE
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=val)
        cell_bg = highlights.get(i, bg) if highlights else bg
        c.fill = _fill(cell_bg); c.border = _border(); c.font = _font()
        if isinstance(val, (int, float)):
            c.alignment = _align("right")
            c.number_format = '#,##0' if val > 99 else '0'
        else:
            c.alignment = _align("left")

def banner(ws, title, sub, ncols):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = title
    c.fill = _fill(GREEN_DARK)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = _align("center")
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]; c.value = sub
    c.fill = _fill(GREEN_MED)
    c.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    c.alignment = _align("center")

def set_widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def total_bar(ws, row, merge_to, label, col_val, value):
    ws.merge_cells(f"A{row}:{get_column_letter(merge_to)}{row}")
    ws[f"A{row}"].value = label
    ws[f"A{row}"].font  = _font(True, WHITE)
    ws[f"A{row}"].fill  = _fill(GREEN_DARK)
    ws[f"A{row}"].alignment = _align("center")
    c = ws.cell(row=row, column=col_val, value=value)
    c.font = _font(True, WHITE); c.fill = _fill(GREEN_DARK)
    c.alignment = _align("right"); c.number_format = '#,##0'
    for col in range(col_val + 1, 20):
        cell = ws.cell(row=row, column=col)
        if cell.value is None:
            cell.fill = _fill(GREEN_DARK)
        else:
            break

def rand_date(year, m_max=12):
    m = random.randint(1, m_max)
    d_max = 28 if m == 2 else (30 if m in [4,6,9,11] else 31)
    return date(year, m, random.randint(1, d_max))


# ════════════════════════════════════════════════════════════════
#  RÉFÉRENTIELS
# ════════════════════════════════════════════════════════════════
DEPOTS = ["Dépôt Korhogo", "Dépôt Bamako"]

PRODUITS = [
    # (code, désignation, catégorie, unité, prix_achat, prix_vente, stock_mini, fournisseur)
    ("PRD-001", "Plants d'anacardier (greffé)",       "Plants fruitiers",       "Plant",  800,  1200, 500, "CNRA Côte d'Ivoire"),
    ("PRD-002", "Plants de manguier (greffé Kent)",   "Plants fruitiers",       "Plant", 1000,  1500, 300, "CNRA Côte d'Ivoire"),
    ("PRD-003", "Plants de papayer Solo",             "Plants fruitiers",       "Plant",  400,   650, 400, "Pépinière Nationale Mali"),
    ("PRD-004", "Plants de bananier plantain",        "Plants fruitiers",       "Rejet",  350,   550, 500, "CNRA Côte d'Ivoire"),
    ("PRD-005", "Plants d'avocatier Hass",            "Plants fruitiers",       "Plant", 1200,  1800, 200, "Semences Agro CI"),
    ("PRD-006", "Plants d'oranger Washington",        "Plants fruitiers",       "Plant",  900,  1400, 200, "Semences Agro CI"),
    ("PRD-007", "Plants de citronnier eureka",        "Plants fruitiers",       "Plant",  750,  1100, 150, "Semences Agro CI"),
    ("PRD-008", "Plants de colatier",                 "Plants fruitiers",       "Plant", 1500,  2200, 100, "CNRA Côte d'Ivoire"),
    ("PRD-009", "Semences de maïs amélioré (1 kg)",  "Céréales & Semences",    "Kg",     800,  1200,  50, "INERA Burkina"),
    ("PRD-010", "Semences de sorgho (1 kg)",          "Céréales & Semences",    "Kg",     600,   900,  40, "INERA Burkina"),
    ("PRD-011", "Semences de mil (1 kg)",             "Céréales & Semences",    "Kg",     550,   850,  40, "Office du Niger"),
    ("PRD-012", "Semences de riz paddy (1 kg)",      "Céréales & Semences",    "Kg",     700,  1050,  60, "Office du Niger"),
    ("PRD-013", "Semences d'arachide (1 kg)",        "Légumineuses",           "Kg",     900,  1350,  30, "INERA Burkina"),
    ("PRD-014", "Semences de niébé (1 kg)",          "Légumineuses",           "Kg",     750,  1100,  30, "INERA Burkina"),
    ("PRD-015", "Semences de soja (1 kg)",           "Légumineuses",           "Kg",     850,  1250,  25, "INERA Burkina"),
    ("PRD-016", "Semences de tomate F1 (sachet 10g)","Cultures maraîchères",   "Sachet",4500,  7000,  50, "Technisem Sénégal"),
    ("PRD-017", "Semences d'oignon Violet de Galmi", "Cultures maraîchères",   "Sachet",3800,  5800,  40, "Technisem Sénégal"),
    ("PRD-018", "Semences de piment (sachet 5g)",    "Cultures maraîchères",   "Sachet",2500,  3800,  30, "Technisem Sénégal"),
    ("PRD-019", "Boutures de manioc (botte 50)",     "Tubercules & Racines",   "Botte",  600,   950, 100, "CNRA Côte d'Ivoire"),
    ("PRD-020", "Semenceaux d'igname Florido (kg)",  "Tubercules & Racines",   "Kg",     450,   700, 150, "CNRA Côte d'Ivoire"),
]

CLIENTS = [
    ("Coopérative Anacardiers Korhogo",  "Korhogo, CI"),
    ("GIE Femmes Rurales Sinématiali",   "Korhogo, CI"),
    ("ANADER – Antenne Korhogo",         "Korhogo, CI"),
    ("Ferme Coulibaly & Fils",           "Korhogo, CI"),
    ("Groupement Maraîchers Ferkessédougou", "Nord CI"),
    ("Office du Niger – Direction Ségou","Ségou, Mali"),
    ("Coopérative OHVN Bamako",          "Bamako, Mali"),
    ("CMDT – Division Agriculture",      "Bamako, Mali"),
    ("Ferme Diallo Horticulture",        "Bamako, Mali"),
    ("GIE Producteurs Manguiers Sikasso","Sikasso, Mali"),
    ("Programme FAO – CI/Mali",          "Abidjan/Bamako"),
    ("UGCPA/BM Burkina",                 "Bobo-Dioulasso"),
]

VENDEURS = [
    ("Koné Drissa",      "Dépôt Korhogo"),
    ("Bamba Mariam",     "Dépôt Korhogo"),
    ("Diallo Seydou",    "Dépôt Bamako"),
    ("Traoré Aminata",   "Dépôt Bamako"),
]

FOURNISSEURS = [
    ("FRN-001","CNRA Côte d'Ivoire",          "Institut public",  "Abidjan – Adiopodoumé",     "27 23 46 00 00", "contact@cnra.ci",         "Plants greffés, boutures, semenceaux"),
    ("FRN-002","INERA Burkina Faso",           "Institut public",  "Bobo-Dioulasso",            "+226 20 97 00 07","inera@inera.bf",          "Semences céréales, légumineuses"),
    ("FRN-003","Office du Niger – Ségou",      "Établissement public","Ségou, Mali",            "+223 21 32 00 12","direction@officeduniger.ml","Semences riz, mil, sorgho"),
    ("FRN-004","Pépinière Nationale du Mali",  "Public",           "Bamako – Sotuba",           "+223 20 22 11 00","pepiniere.mali@gov.ml",   "Plants fruitiers, arbres forestiers"),
    ("FRN-005","Semences Agro CI (SACI)",      "Privé",            "Yamoussoukro",              "27 30 64 00 00", "saci@saci.ci",            "Plants fruitiers greffés premium"),
    ("FRN-006","Technisem Sénégal",            "Privé",            "Dakar – Zone Franche",      "+221 33 820 45 00","technisem@technisem.com","Semences maraîchères F1 importées"),
    ("FRN-007","GIZ – Programme Semences",     "ONG/Coopération",  "Ouagadougou / Bamako",      "+226 25 30 67 00","giz.seeds@giz.de",       "Semences améliorées, formations"),
    ("FRN-008","UGCPA Bobo-Dioulasso",         "Coopérative",      "Bobo-Dioulasso, Burkina",   "+226 20 98 12 34","ugcpa@ugcpa.org",         "Arachide, niébé, soja certifiés"),
]

MODES_PAI   = ["ESPÈCES", "MOBILE MONEY", "VIREMENT BANCAIRE", "BON DE COMMANDE"]
STAT_VENTE  = ["ENCAISSÉE","ENCAISSÉE","ENCAISSÉE","CRÉDIT"]
STAT_CMD    = ["LIVRÉE","LIVRÉE","LIVRÉE","EN COURS","ANNULÉE"]
STAT_APPRO  = ["RÉCEPTIONNÉ","RÉCEPTIONNÉ","RÉCEPTIONNÉ","PARTIEL","EN ATTENTE"]


# ════════════════════════════════════════════════════════════════
#  RAPPORT VENTES
# ════════════════════════════════════════════════════════════════
def rapport_ventes():
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    MOIS = ["Janvier","Février","Mars","Avril","Mai","Juin",
            "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    for year, n_ventes, m_max in [(2025, 260, 12), (2026, 110, 5)]:
        # ── Détail ──
        ws = wb.create_sheet(f"Ventes {year}")
        ws.freeze_panes = "A5"
        cols = ["N° Vente","Date","Client","Dépôt","Produit","Catégorie",
                "Qté","Prix unit. (FCFA)","Total (FCFA)","Mode paiement","Statut","Vendeur"]
        banner(ws, f"RAPPORT DÉTAILLÉ DES VENTES – {year}",
               f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
               len(cols))
        header_row(ws, 4, cols)
        set_widths(ws, [13,11,36,16,32,20,7,18,18,20,12,20])

        ventes = []
        for i in range(1, n_ventes + 1):
            d   = rand_date(year, m_max)
            cl, loc = random.choice(CLIENTS)
            p   = random.choice(PRODUITS)
            qty = random.randint(10, 500)
            tot = qty * p[5]
            mod = random.choice(MODES_PAI)
            sta = random.choice(STAT_VENTE)
            vend, depot = random.choice(VENDEURS)
            ventes.append((f"VTE-{year}-{i:04d}", d, cl, depot,
                           p[1], p[2], qty, p[5], tot, mod, sta, vend))
        ventes.sort(key=lambda x: x[1])

        totaux_m = {}
        nb_m     = {}
        for idx, v in enumerate(ventes):
            data_row(ws, 5+idx, v, alt=(idx%2==1))
            ws.row_dimensions[5+idx].height = 18
            m = v[1].month
            totaux_m[m] = totaux_m.get(m, 0) + v[8]
            nb_m[m]     = nb_m.get(m, 0) + 1

        total_global = sum(v[8] for v in ventes)
        total_bar(ws, 5+len(ventes), 8, "TOTAL GÉNÉRAL", 9, total_global)

        # ── Résumé mensuel ──
        ws2 = wb.create_sheet(f"Résumé {year}")
        ws2.freeze_panes = "A5"
        banner(ws2, f"RÉSUMÉ MENSUEL – {year}",
               f"AGSV – Consolidation  |  Généré le {date.today():%d/%m/%Y}", 4)
        header_row(ws2, 4, ["Mois","Nb ventes","CA (FCFA)","% du CA annuel"])
        set_widths(ws2, [16,14,22,18])

        for idx, m in enumerate(range(1, m_max+1)):
            ca  = totaux_m.get(m, 0)
            nb  = nb_m.get(m, 0)
            pct = f"{round(ca/total_global*100,1)} %" if total_global else "0 %"
            data_row(ws2, 5+idx, [MOIS[m-1], nb, ca, pct], alt=(idx%2==1))
            ws2.row_dimensions[5+idx].height = 18

        tr2 = 5 + m_max
        total_bar(ws2, tr2, 2, "TOTAL", 3, total_global)
        ws2.cell(row=tr2, column=4).value = "100 %"
        ws2.cell(row=tr2, column=4).fill  = _fill(GREEN_DARK)
        ws2.cell(row=tr2, column=4).font  = _font(True, WHITE)
        ws2.cell(row=tr2, column=4).alignment = _align("center")

    path = "/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Ventes_2025-2026.xlsx"
    wb.save(path)
    print("✓ Rapport_Ventes_2025-2026.xlsx")


# ════════════════════════════════════════════════════════════════
#  RAPPORT COMMANDES
# ════════════════════════════════════════════════════════════════
def rapport_commandes():
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    for year, n, m_max in [(2025, 170, 12), (2026, 80, 5)]:
        ws = wb.create_sheet(f"Commandes {year}")
        ws.freeze_panes = "A5"
        cols = ["N° Commande","Date commande","Livraison prévue","Client","Dépôt",
                "Produit","Catégorie","Qté commandée","Qté livrée",
                "Prix unit. (FCFA)","Montant livré (FCFA)","Statut","Vendeur"]
        banner(ws, f"RAPPORT DES COMMANDES – {year}",
               f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
               len(cols))
        header_row(ws, 4, cols)
        set_widths(ws, [14,13,16,36,16,32,20,14,11,18,20,12,20])

        cmds = []
        for i in range(1, n+1):
            d_cmd  = rand_date(year, m_max)
            d_livr = d_cmd + timedelta(days=random.randint(3, 21))
            cl, _  = random.choice(CLIENTS)
            p      = random.choice(PRODUITS)
            _, depot = random.choice(VENDEURS)
            q_cmd  = random.randint(20, 800)
            stat   = random.choice(STAT_CMD)
            q_liv  = q_cmd if stat=="LIVRÉE" else (random.randint(0,q_cmd-1) if stat=="EN COURS" else 0)
            mont   = q_liv * p[5]
            vend, depot = random.choice(VENDEURS)
            cmds.append((f"CMD-{year}-{i:04d}", d_cmd, d_livr, cl, depot,
                         p[1], p[2], q_cmd, q_liv, p[5], mont, stat, vend))
        cmds.sort(key=lambda x: x[1])

        total_global = sum(c[10] for c in cmds)
        for idx, v in enumerate(cmds):
            hl = {12: GREEN_LIGHT if v[11]=="LIVRÉE" else (RED_LIGHT if v[11]=="ANNULÉE" else YELLOW_L)}
            data_row(ws, 5+idx, v, alt=(idx%2==1), highlights=hl)
            ws.row_dimensions[5+idx].height = 18

        total_bar(ws, 5+len(cmds), 10, "TOTAL LIVRÉ", 11, total_global)

        # Résumé par statut
        ws2 = wb.create_sheet(f"Statuts {year}")
        banner(ws2, f"COMMANDES PAR STATUT – {year}",
               f"AGSV  |  Généré le {date.today():%d/%m/%Y}", 4)
        header_row(ws2, 4, ["Statut","Nb commandes","Montant (FCFA)","Part (%)"])
        set_widths(ws2, [18,16,24,14])
        stat_map = {}
        for c in cmds:
            s = c[11]
            stat_map.setdefault(s, [0, 0])
            stat_map[s][0] += 1; stat_map[s][1] += c[10]
        for idx, (s, (nb, mt)) in enumerate(sorted(stat_map.items())):
            pct = f"{round(mt/total_global*100,1)} %" if total_global else "0 %"
            hl  = {1: GREEN_LIGHT if s=="LIVRÉE" else (RED_LIGHT if s=="ANNULÉE" else YELLOW_L)}
            data_row(ws2, 5+idx, [s, nb, mt, pct], alt=(idx%2==1), highlights=hl)
            ws2.row_dimensions[5+idx].height = 18

    path = "/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Commandes_2025-2026.xlsx"
    wb.save(path)
    print("✓ Rapport_Commandes_2025-2026.xlsx")


# ════════════════════════════════════════════════════════════════
#  RAPPORT FOURNISSEURS
# ════════════════════════════════════════════════════════════════
def rapport_fournisseurs():
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # Répertoire
    ws = wb.create_sheet("Fournisseurs")
    ws.freeze_panes = "A5"
    cols = ["Code","Nom fournisseur","Type","Adresse",
            "Téléphone","Email","Produits fournis",
            "Délai livr. (j)","Conditions paiement","Statut"]
    banner(ws, "RÉPERTOIRE DES FOURNISSEURS",
           f"AGSV – Partenaires agricoles  |  Mis à jour le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    set_widths(ws, [10,28,18,28,20,28,36,14,24,10])

    CONDITIONS = ["Paiement à la livraison","30 jours net","15 jours net","Contre remboursement"]
    DELAIS     = [5, 7, 10, 14, 21]
    for idx, f in enumerate(FOURNISSEURS):
        code, nom, typ, adr, tel, mail, produits = f
        cond   = random.choice(CONDITIONS)
        delai  = random.choice(DELAIS)
        statut = "ACTIF"
        hl = {10: GREEN_LIGHT}
        data_row(ws, 5+idx, [code, nom, typ, adr, tel, mail, produits, delai, cond, statut],
                 alt=(idx%2==1), highlights=hl)
        ws.row_dimensions[5+idx].height = 22

    # Historique approvisionnements
    ws2 = wb.create_sheet("Approvisionnements")
    ws2.freeze_panes = "A5"
    cols2 = ["N° Appro","Date","Fournisseur","Produit","Catégorie",
             "Qté reçue","Unité","Prix unit. (FCFA)","Montant (FCFA)",
             "Dépôt récepteur","Statut","N° Bon commande"]
    banner(ws2, "HISTORIQUE DES APPROVISIONNEMENTS – 2024 / 2025 / 2026",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols2))
    header_row(ws2, 4, cols2)
    set_widths(ws2, [13,11,26,30,20,10,8,18,20,18,14,16])

    appros = []
    for year in [2024, 2025, 2026]:
        n = 35 if year==2024 else (55 if year==2025 else 18)
        for i in range(n):
            d  = rand_date(year, 5 if year==2026 else 12)
            f  = random.choice(FOURNISSEURS)
            p  = random.choice(PRODUITS)
            qty = random.randint(50, 2000)
            pu  = int(p[4] * random.uniform(0.85, 0.97))
            tot = qty * pu
            dep = random.choice(DEPOTS)
            sta = random.choice(STAT_APPRO)
            appros.append((f"APP-{year}-{i+1:04d}", d, f[1], p[1], p[2],
                           qty, p[3], pu, tot, dep, sta, f"BC-{year}-{i+1:04d}"))

    appros.sort(key=lambda x: x[1])
    total_g = sum(a[8] for a in appros)
    for idx, v in enumerate(appros):
        hl = {11: GREEN_LIGHT if v[10]=="RÉCEPTIONNÉ" else (YELLOW_L if v[10]=="PARTIEL" else RED_LIGHT)}
        data_row(ws2, 5+idx, v, alt=(idx%2==1), highlights=hl)
        ws2.row_dimensions[5+idx].height = 18

    total_bar(ws2, 5+len(appros), 8, "TOTAL GÉNÉRAL", 9, total_g)

    path = "/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Fournisseurs.xlsx"
    wb.save(path)
    print("✓ Rapport_Fournisseurs.xlsx")


# ════════════════════════════════════════════════════════════════
#  RAPPORT PRODUITS
# ════════════════════════════════════════════════════════════════
def rapport_produits():
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    stock_par_depot = {}
    for p in PRODUITS:
        stock_par_depot[p[0]] = {
            "Dépôt Korhogo": random.randint(p[6], p[6]*6),
            "Dépôt Bamako":  random.randint(p[6]//2, p[6]*4),
        }

    # ── Catalogue ──
    ws = wb.create_sheet("Catalogue Produits")
    ws.freeze_panes = "A5"
    cols = ["Code","Désignation","Catégorie","Unité",
            "Prix achat (FCFA)","Prix vente (FCFA)","Marge (FCFA)","Marge (%)",
            "Stock mini","Fournisseur principal","Statut"]
    banner(ws, "CATALOGUE DES PRODUITS AGRICOLES",
           f"AGSV – Référentiel produits  |  Mis à jour le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    set_widths(ws, [10,32,22,8,18,18,14,10,10,24,10])

    for idx, p in enumerate(PRODUITS):
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        marge   = p_vte - p_ach
        marge_p = round(marge/p_ach*100, 1)
        stk     = sum(stock_par_depot[code].values())
        statut  = "ACTIF" if stk > s_min else "RUPTURE"
        hl = {11: GREEN_LIGHT if statut=="ACTIF" else RED_LIGHT}
        data_row(ws, 5+idx,
                 [code, nom, cat, unit, p_ach, p_vte, marge, f"{marge_p} %", s_min, fourn, statut],
                 alt=(idx%2==1), highlights=hl)
        ws.row_dimensions[5+idx].height = 20

    # ── Stock par dépôt ──
    ws2 = wb.create_sheet("Stock par Dépôt")
    ws2.freeze_panes = "A5"
    cols2 = ["Code","Désignation","Catégorie","Unité",
             "Dépôt Korhogo","Dépôt Bamako","Stock total",
             "Valeur stock (FCFA)","Stock mini","Alerte"]
    banner(ws2, "ÉTAT DES STOCKS PAR DÉPÔT",
           f"AGSV – Situation au {date.today():%d/%m/%Y}", len(cols2))
    header_row(ws2, 4, cols2)
    set_widths(ws2, [10,32,22,8,16,14,12,22,10,14])

    val_total = 0
    for idx, p in enumerate(PRODUITS):
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        dk = stock_par_depot[code]["Dépôt Korhogo"]
        db = stock_par_depot[code]["Dépôt Bamako"]
        tot = dk + db
        val = tot * p_vte
        val_total += val
        alerte = "⚠ STOCK BAS" if tot <= s_min*1.5 else ("✓ OK" if tot > s_min*3 else "→ MOYEN")
        hl = {10: RED_LIGHT if "⚠" in alerte else (GREEN_LIGHT if "✓" in alerte else YELLOW_L)}
        data_row(ws2, 5+idx, [code, nom, cat, unit, dk, db, tot, val, s_min, alerte],
                 alt=(idx%2==1), highlights=hl)
        ws2.row_dimensions[5+idx].height = 20

    total_bar(ws2, 5+len(PRODUITS), 7, "VALEUR TOTALE DU STOCK", 8, val_total)

    # ── Mouvements ──
    ws3 = wb.create_sheet("Mouvements de Stock")
    ws3.freeze_panes = "A5"
    cols3 = ["N° Mouvement","Date","Produit","Catégorie","Dépôt",
             "Type","Quantité","Unité","Motif","Référence"]
    banner(ws3, "HISTORIQUE DES MOUVEMENTS DE STOCK – 2025/2026",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols3))
    header_row(ws3, 4, cols3)
    set_widths(ws3, [14,11,32,22,16,12,10,8,30,14])

    MOTIFS = {
        "ENTRÉE":    ["Approvisionnement fournisseur","Retour client","Transfert inter-dépôt"],
        "SORTIE":    ["Vente client","Vente client","Vente client","Transfert inter-dépôt","Don/subvention"],
        "PERTE":     ["Plantes mortes (transport)","Semences avariées","Détérioration stockage"],
        "INVENTAIRE":["Inventaire mensuel","Inventaire trimestriel","Ajustement inventaire"],
    }
    TYPES  = ["ENTRÉE","SORTIE","SORTIE","SORTIE","PERTE","INVENTAIRE"]
    TYPE_C = {"ENTRÉE":GREEN_LIGHT,"SORTIE":WHITE,"PERTE":RED_LIGHT,"INVENTAIRE":YELLOW_L}

    mvts = []
    for year in [2025, 2026]:
        n = 280 if year==2025 else 110
        for i in range(n):
            d   = rand_date(year, 5 if year==2026 else 12)
            p   = random.choice(PRODUITS)
            dep = random.choice(DEPOTS)
            typ = random.choice(TYPES)
            qty = random.randint(5, 300)
            mot = random.choice(MOTIFS[typ])
            ref = f"MVT-{year}-{i+1:04d}"
            mvts.append((ref, d, p[1], p[2], dep, typ, qty, p[3], mot, ref.replace("MVT","REF")))

    mvts.sort(key=lambda x: x[1])
    for idx, v in enumerate(mvts):
        hl = {6: TYPE_C.get(v[5], WHITE)}
        data_row(ws3, 5+idx, v, alt=(idx%2==1), highlights=hl)
        ws3.row_dimensions[5+idx].height = 18

    # ── Performance ──
    ws4 = wb.create_sheet("Performance Produits")
    ws4.freeze_panes = "A5"
    cols4 = ["Code","Désignation","Catégorie",
             "Qté vendue 2025","CA 2025 (FCFA)","Qté vendue 2026 (jan–mai)","CA 2026 (FCFA)",
             "Évolution CA","Rang 2025"]
    banner(ws4, "PERFORMANCE DES VENTES PAR PRODUIT – 2025 vs 2026",
           f"AGSV – Analyse comparative  |  Généré le {date.today():%d/%m/%Y}", len(cols4))
    header_row(ws4, 4, cols4)
    set_widths(ws4, [10,32,22,16,20,22,20,14,10])

    perfs = []
    for p in PRODUITS:
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        q25  = random.randint(200, 3000)
        ca25 = q25 * p_vte
        q26  = random.randint(80, int(q25*0.5))
        ca26 = q26 * p_vte
        annl = ca26 * 12/5
        evol = round((annl - ca25)/ca25*100, 1)
        perfs.append((code, nom, cat, q25, ca25, q26, ca26, evol))

    perfs.sort(key=lambda x: -x[4])
    for idx, (code, nom, cat, q25, ca25, q26, ca26, evol) in enumerate(perfs):
        evol_str = f"▲ +{evol} %" if evol >= 0 else f"▼ {evol} %"
        hl = {8: GREEN_LIGHT if evol >= 0 else RED_LIGHT}
        data_row(ws4, 5+idx, [code, nom, cat, q25, ca25, q26, ca26, evol_str, idx+1],
                 alt=(idx%2==1), highlights=hl)
        ws4.row_dimensions[5+idx].height = 20

    tr4 = 5+len(perfs)
    total_bar(ws4, tr4, 3, "TOTAUX", 4, sum(x[3] for x in perfs))
    for col_i, val in enumerate([sum(x[4] for x in perfs),
                                  sum(x[5] for x in perfs),
                                  sum(x[6] for x in perfs)], start=5):
        c = ws4.cell(row=tr4, column=col_i, value=val)
        c.font = _font(True, WHITE); c.fill = _fill(GREEN_DARK)
        c.alignment = _align("right"); c.number_format = '#,##0'

    # ── Résumé par catégorie ──
    ws5 = wb.create_sheet("Résumé par Catégorie")
    ws5.freeze_panes = "A5"
    cols5 = ["Catégorie","Nb références","Valeur stock (FCFA)",
             "CA 2025 (FCFA)","CA 2026 jan–mai (FCFA)","Marge moyenne (%)"]
    banner(ws5, "ANALYSE PAR CATÉGORIE DE PRODUITS AGRICOLES",
           f"AGSV – Synthèse  |  Généré le {date.today():%d/%m/%Y}", len(cols5))
    header_row(ws5, 4, cols5)
    set_widths(ws5, [26,16,24,24,26,18])

    cats = {}
    for p in PRODUITS:
        code, nom, cat, unit, p_ach, p_vte, s_min, fourn = p
        stk   = sum(stock_par_depot[code].values()) * p_vte
        marge = round((p_vte-p_ach)/p_ach*100, 1)
        cats.setdefault(cat, {"nb":0,"val":0,"ca25":0,"ca26":0,"marges":[]})
        cats[cat]["nb"]    += 1
        cats[cat]["val"]   += stk
        cats[cat]["ca25"]  += random.randint(500000, 5000000)
        cats[cat]["ca26"]  += random.randint(200000, 2000000)
        cats[cat]["marges"].append(marge)

    for idx, (cat, d) in enumerate(sorted(cats.items())):
        moy = round(sum(d["marges"])/len(d["marges"]), 1)
        data_row(ws5, 5+idx, [cat, d["nb"], d["val"], d["ca25"], d["ca26"], f"{moy} %"],
                 alt=(idx%2==1))
        ws5.row_dimensions[5+idx].height = 22

    path = "/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Produits.xlsx"
    wb.save(path)
    print("✓ Rapport_Produits.xlsx")


# ── Run ───────────────────────────────────────────────────────
rapport_ventes()
rapport_commandes()
rapport_fournisseurs()
rapport_produits()
print("\nTous les fichiers sont dans /rapports/")