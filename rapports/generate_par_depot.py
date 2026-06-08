import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import date, timedelta

random.seed(77)

GREEN_DARK  = "1A6B3A"
GREEN_MED   = "2E7D32"
GREEN_LIGHT = "E8F5E9"
WHITE       = "FFFFFF"
GREY_ROW    = "FAFAFA"
RED_LIGHT   = "FFEBEE"
YELLOW_L    = "FFFDE7"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(bold=False, color="1A1A1A", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    t = Side(style="thin", color="DDDDDD")
    return Border(left=t, right=t, top=t, bottom=t)
def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def header_row(ws, row, cols):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.fill = _fill(GREEN_DARK); c.font = _font(True, WHITE, 10)
        c.alignment = _align("center"); c.border = _border()

def data_row(ws, row, vals, alt=False, hl=None):
    bg = GREY_ROW if alt else WHITE
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.fill = _fill(hl.get(i, bg) if hl else bg)
        c.border = _border(); c.font = _font()
        if isinstance(val, (int, float)):
            c.alignment = _align("right")
            c.number_format = '#,##0' if val > 99 else '0'
        else:
            c.alignment = _align("left")

def banner(ws, title, sub, ncols):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c = ws["A1"]; c.value = title
    c.fill = _fill(GREEN_DARK)
    c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
    c.alignment = _align("center")
    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    c = ws["A2"]; c.value = sub
    c.fill = _fill(GREEN_MED)
    c.font = Font(italic=True, color=WHITE, size=10, name="Calibri")
    c.alignment = _align("center")

def set_widths(ws, wlist):
    for i, w in enumerate(wlist, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def total_bar(ws, row, merge_to, label, col_val, value):
    ws.merge_cells(f"A{row}:{get_column_letter(merge_to)}{row}")
    ws[f"A{row}"].value = label
    ws[f"A{row}"].font  = _font(True, WHITE)
    ws[f"A{row}"].fill  = _fill(GREEN_DARK)
    ws[f"A{row}"].alignment = _align("center")
    c = ws.cell(row=row, column=col_val, value=value)
    c.font = _font(True, WHITE); c.fill = _fill(GREEN_DARK)
    c.alignment = _align("right"); c.number_format = '#,##0'

def rand_date(year, m_max=12):
    m = random.randint(1, m_max)
    d_max = 28 if m == 2 else (30 if m in [4,6,9,11] else 31)
    return date(year, m, random.randint(1, d_max))


# ════════════════════════════════════════════════════════════════
#  RÉFÉRENTIELS — PAR DÉPÔT
# ════════════════════════════════════════════════════════════════

DEPOTS = {
    "Korhogo": {
        "nom":    "Dépôt Korhogo",
        "ville":  "Korhogo, Côte d'Ivoire",
        "vendeurs": ["Koné Drissa", "Bamba Mariam"],
        "produits": [
            ("PRD-K01","Plants d'anacardier (greffé)",       "Plants fruitiers",    "Plant",  800, 1200, 500,"CNRA CI"),
            ("PRD-K02","Plants de manguier Kent (greffé)",   "Plants fruitiers",    "Plant", 1000, 1500, 300,"CNRA CI"),
            ("PRD-K03","Plants de papayer Solo",             "Plants fruitiers",    "Plant",  400,  650, 400,"CNRA CI"),
            ("PRD-K04","Plants de bananier plantain",        "Plants fruitiers",    "Rejet",  350,  550, 500,"CNRA CI"),
            ("PRD-K05","Plants d'avocatier Hass",            "Plants fruitiers",    "Plant", 1200, 1800, 200,"Semences Agro CI"),
            ("PRD-K06","Plants de colatier",                 "Plants fruitiers",    "Plant", 1500, 2200, 100,"CNRA CI"),
            ("PRD-K07","Semences de maïs amélioré (1 kg)",  "Céréales & Semences", "Kg",     800, 1200,  80,"INERA Burkina"),
            ("PRD-K08","Semenceaux d'igname Florido (kg)",  "Tubercules & Racines","Kg",     450,  700, 200,"CNRA CI"),
            ("PRD-K09","Boutures de manioc (botte 50)",     "Tubercules & Racines","Botte",  600,  950, 150,"CNRA CI"),
            ("PRD-K10","Semences d'arachide (1 kg)",        "Légumineuses",        "Kg",     900, 1350,  60,"INERA Burkina"),
        ],
        "clients": [
            "Coopérative Anacardiers Korhogo",
            "GIE Femmes Rurales Sinématiali",
            "ANADER – Antenne Korhogo",
            "Ferme Coulibaly & Fils",
            "Groupement Maraîchers Ferkessédougou",
            "Coopérative Fruitiers Boundiali",
            "Union des Producteurs du Poro",
            "Programme FIRCA Korhogo",
        ],
        "fournisseurs": [
            ("FRN-K01","CNRA Côte d'Ivoire","Institut public","Abidjan – Adiopodoumé","27 23 46 00 00","contact@cnra.ci","Plants greffés, boutures, semenceaux","Paiement à la livraison",7),
            ("FRN-K02","Semences Agro CI","Privé","Yamoussoukro","27 30 64 00 00","saci@saci.ci","Plants fruitiers premium","30 jours net",10),
            ("FRN-K03","INERA Burkina Faso","Institut public","Bobo-Dioulasso","+226 20 97 00 07","inera@inera.bf","Semences céréales, légumineuses","15 jours net",14),
            ("FRN-K04","GIZ – Programme Semences","ONG","Abidjan/Ouaga","+226 25 30 67 00","giz.seeds@giz.de","Semences améliorées certifiées","Contra remboursement",21),
            ("FRN-K05","FIRCA CI","Fonds interprofessionnel","Abidjan – Plateau","27 20 25 00 00","firca@firca.ci","Intrants agricoles subventionnés","BON DE COMMANDE",5),
        ],
    },

    "Bamako": {
        "nom":    "Dépôt Bamako",
        "ville":  "Bamako, Mali",
        "vendeurs": ["Diallo Seydou", "Traoré Aminata"],
        "produits": [
            ("PRD-B01","Semences de mil (1 kg)",             "Céréales & Semences", "Kg",     550,  850,  80,"Office du Niger"),
            ("PRD-B02","Semences de sorgho (1 kg)",          "Céréales & Semences", "Kg",     600,  900,  80,"Office du Niger"),
            ("PRD-B03","Semences de riz paddy (1 kg)",      "Céréales & Semences", "Kg",     700, 1050, 100,"Office du Niger"),
            ("PRD-B04","Semences de maïs amélioré (1 kg)",  "Céréales & Semences", "Kg",     800, 1200,  60,"INERA Burkina"),
            ("PRD-B05","Semences de niébé (1 kg)",          "Légumineuses",        "Kg",     750, 1100,  50,"UGCPA Bobo"),
            ("PRD-B06","Semences de soja (1 kg)",           "Légumineuses",        "Kg",     850, 1250,  40,"UGCPA Bobo"),
            ("PRD-B07","Semences d'oignon Violet de Galmi","Cultures maraîchères", "Sachet",3800, 5800,  60,"Technisem Sénégal"),
            ("PRD-B08","Semences de tomate F1 (sachet 10g)","Cultures maraîchères","Sachet",4500, 7000,  50,"Technisem Sénégal"),
            ("PRD-B09","Semences de piment (sachet 5g)",    "Cultures maraîchères","Sachet",2500, 3800,  40,"Technisem Sénégal"),
            ("PRD-B10","Plants de manguier Amélie",         "Plants fruitiers",    "Plant",  900, 1350, 200,"Pépinière Nat. Mali"),
        ],
        "clients": [
            "Office du Niger – Direction Ségou",
            "Coopérative OHVN Bamako",
            "CMDT – Division Agriculture",
            "Ferme Diallo Horticulture",
            "GIE Producteurs Manguiers Sikasso",
            "Association Maraîchers Baguinéda",
            "Programme PNUD Agriculture Mali",
            "Groupement Femmes Rurales Koulikoro",
        ],
        "fournisseurs": [
            ("FRN-B01","Office du Niger – Ségou","Établissement public","Ségou, Mali","+223 21 32 00 12","direction@officeduniger.ml","Semences riz, mil, sorgho","Paiement à la livraison",7),
            ("FRN-B02","Pépinière Nationale du Mali","Public","Bamako – Sotuba","+223 20 22 11 00","pepiniere.mali@gov.ml","Plants fruitiers, arbres","30 jours net",10),
            ("FRN-B03","Technisem Sénégal","Privé","Dakar – Zone Franche","+221 33 820 45 00","technisem@technisem.com","Semences maraîchères F1","Contre remboursement",14),
            ("FRN-B04","UGCPA Bobo-Dioulasso","Coopérative","Bobo-Dioulasso, Burkina","+226 20 98 12 34","ugcpa@ugcpa.org","Niébé, soja certifiés","15 jours net",10),
            ("FRN-B05","INERA Burkina Faso","Institut public","Bobo-Dioulasso","+226 20 97 00 07","inera@inera.bf","Semences céréales améliorées","15 jours net",14),
        ],
    },
}

MOIS = ["Janvier","Février","Mars","Avril","Mai","Juin",
        "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
MODES_PAI   = ["ESPÈCES","MOBILE MONEY","VIREMENT BANCAIRE","BON DE COMMANDE"]
STAT_VENTE  = ["ENCAISSÉE","ENCAISSÉE","ENCAISSÉE","CRÉDIT"]
STAT_CMD    = ["LIVRÉE","LIVRÉE","LIVRÉE","EN COURS","ANNULÉE"]
STAT_APPRO  = ["RÉCEPTIONNÉ","RÉCEPTIONNÉ","RÉCEPTIONNÉ","PARTIEL","EN ATTENTE"]


# ════════════════════════════════════════════════════════════════
#  VENTES
# ════════════════════════════════════════════════════════════════
def generer_ventes(depot_key):
    d   = DEPOTS[depot_key]
    wb  = openpyxl.Workbook(); wb.remove(wb.active)

    for year, n, m_max in [(2025, 200, 12), (2026, 90, 5)]:
        ws = wb.create_sheet(f"Ventes {year}")
        ws.freeze_panes = "A5"
        cols = ["N° Vente","Date","Client","Produit","Catégorie",
                "Qté","Unité","Prix unit. (FCFA)","Total (FCFA)","Mode paiement","Statut","Vendeur"]
        banner(ws, f"RAPPORT DES VENTES – {year}  |  {d['nom'].upper()}",
               f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
               len(cols))
        header_row(ws, 4, cols)
        set_widths(ws, [13,11,36,32,22,7,7,18,18,20,12,20])

        ventes = []
        for i in range(1, n+1):
            dt  = rand_date(year, m_max)
            cl  = random.choice(d["clients"])
            p   = random.choice(d["produits"])
            qty = random.randint(10, 600)
            tot = qty * p[5]
            mod = random.choice(MODES_PAI)
            sta = random.choice(STAT_VENTE)
            vnd = random.choice(d["vendeurs"])
            ventes.append((f"VTE-{depot_key[:3].upper()}-{year}-{i:04d}",
                           dt, cl, p[1], p[2], qty, p[3], p[5], tot, mod, sta, vnd))
        ventes.sort(key=lambda x: x[1])

        totaux_m = {}; nb_m = {}
        for idx, v in enumerate(ventes):
            data_row(ws, 5+idx, v, alt=(idx%2==1))
            ws.row_dimensions[5+idx].height = 18
            m = v[1].month
            totaux_m[m] = totaux_m.get(m,0) + v[8]
            nb_m[m]     = nb_m.get(m,0) + 1

        total_g = sum(v[8] for v in ventes)
        total_bar(ws, 5+len(ventes), 8, "TOTAL GÉNÉRAL", 9, total_g)

        # Résumé mensuel
        ws2 = wb.create_sheet(f"Résumé {year}")
        banner(ws2, f"RÉSUMÉ MENSUEL – {year}  |  {d['nom'].upper()}",
               f"AGSV  |  Généré le {date.today():%d/%m/%Y}", 4)
        header_row(ws2, 4, ["Mois","Nb ventes","CA (FCFA)","% du CA annuel"])
        set_widths(ws2, [16,14,22,18])
        for idx, m in enumerate(range(1, m_max+1)):
            ca  = totaux_m.get(m,0)
            nb  = nb_m.get(m,0)
            pct = f"{round(ca/total_g*100,1)} %" if total_g else "0 %"
            data_row(ws2, 5+idx, [MOIS[m-1], nb, ca, pct], alt=(idx%2==1))
            ws2.row_dimensions[5+idx].height = 18
        total_bar(ws2, 5+m_max, 2, "TOTAL", 3, total_g)

    path = f"/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Ventes_{depot_key}_2025-2026.xlsx"
    wb.save(path)
    print(f"✓ Rapport_Ventes_{depot_key}_2025-2026.xlsx")


# ════════════════════════════════════════════════════════════════
#  COMMANDES
# ════════════════════════════════════════════════════════════════
def generer_commandes(depot_key):
    d  = DEPOTS[depot_key]
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    for year, n, m_max in [(2025, 140, 12), (2026, 60, 5)]:
        ws = wb.create_sheet(f"Commandes {year}")
        ws.freeze_panes = "A5"
        cols = ["N° Commande","Date commande","Livraison prévue","Client",
                "Produit","Catégorie","Unité","Qté commandée","Qté livrée",
                "Prix unit. (FCFA)","Montant livré (FCFA)","Statut","Vendeur"]
        banner(ws, f"RAPPORT DES COMMANDES – {year}  |  {d['nom'].upper()}",
               f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
               len(cols))
        header_row(ws, 4, cols)
        set_widths(ws, [16,13,16,36,32,22,7,14,11,18,20,12,20])

        cmds = []
        for i in range(1, n+1):
            d_cmd  = rand_date(year, m_max)
            d_livr = d_cmd + timedelta(days=random.randint(3,21))
            cl     = random.choice(d["clients"])
            p      = random.choice(d["produits"])
            q_cmd  = random.randint(20, 1000)
            stat   = random.choice(STAT_CMD)
            q_liv  = q_cmd if stat=="LIVRÉE" else (random.randint(0,q_cmd-1) if stat=="EN COURS" else 0)
            mont   = q_liv * p[5]
            vnd    = random.choice(d["vendeurs"])
            cmds.append((f"CMD-{depot_key[:3].upper()}-{year}-{i:04d}",
                         d_cmd, d_livr, cl, p[1], p[2], p[3],
                         q_cmd, q_liv, p[5], mont, stat, vnd))
        cmds.sort(key=lambda x: x[1])

        total_g = sum(c[10] for c in cmds)
        for idx, v in enumerate(cmds):
            h = {12: GREEN_LIGHT if v[11]=="LIVRÉE" else (RED_LIGHT if v[11]=="ANNULÉE" else YELLOW_L)}
            data_row(ws, 5+idx, v, alt=(idx%2==1), hl=h)
            ws.row_dimensions[5+idx].height = 18
        total_bar(ws, 5+len(cmds), 10, "TOTAL LIVRÉ", 11, total_g)

        # Résumé statuts
        ws2 = wb.create_sheet(f"Statuts {year}")
        banner(ws2, f"COMMANDES PAR STATUT – {year}  |  {d['nom'].upper()}",
               f"AGSV  |  Généré le {date.today():%d/%m/%Y}", 4)
        header_row(ws2, 4, ["Statut","Nb commandes","Montant (FCFA)","Part (%)"])
        set_widths(ws2, [18,16,24,14])
        stat_map = {}
        for c in cmds:
            stat_map.setdefault(c[11],[0,0])
            stat_map[c[11]][0] += 1; stat_map[c[11]][1] += c[10]
        for idx, (s,(nb,mt)) in enumerate(sorted(stat_map.items())):
            pct = f"{round(mt/total_g*100,1)} %" if total_g else "0 %"
            h   = {1: GREEN_LIGHT if s=="LIVRÉE" else (RED_LIGHT if s=="ANNULÉE" else YELLOW_L)}
            data_row(ws2, 5+idx, [s, nb, mt, pct], alt=(idx%2==1), hl=h)
            ws2.row_dimensions[5+idx].height = 18

    path = f"/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Commandes_{depot_key}_2025-2026.xlsx"
    wb.save(path)
    print(f"✓ Rapport_Commandes_{depot_key}_2025-2026.xlsx")


# ════════════════════════════════════════════════════════════════
#  FOURNISSEURS
# ════════════════════════════════════════════════════════════════
def generer_fournisseurs(depot_key):
    d  = DEPOTS[depot_key]
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # Répertoire
    ws = wb.create_sheet("Fournisseurs")
    ws.freeze_panes = "A5"
    cols = ["Code","Nom fournisseur","Type","Adresse",
            "Téléphone","Email","Produits fournis","Délai livr. (j)","Conditions paiement","Statut"]
    banner(ws, f"RÉPERTOIRE DES FOURNISSEURS  |  {d['nom'].upper()}",
           f"AGSV – Partenaires agricoles  |  Mis à jour le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    set_widths(ws, [10,28,18,30,20,30,34,14,26,10])

    for idx, f in enumerate(d["fournisseurs"]):
        code, nom, typ, adr, tel, mail, produits, cond, delai = f
        data_row(ws, 5+idx,
                 [code,nom,typ,adr,tel,mail,produits,delai,cond,"ACTIF"],
                 alt=(idx%2==1), hl={10: GREEN_LIGHT})
        ws.row_dimensions[5+idx].height = 22

    # Historique approvisionnements
    ws2 = wb.create_sheet("Approvisionnements")
    ws2.freeze_panes = "A5"
    cols2 = ["N° Appro","Date","Fournisseur","Produit","Catégorie",
             "Qté reçue","Unité","Prix unit. (FCFA)","Montant (FCFA)","Statut","N° Bon commande"]
    banner(ws2, f"APPROVISIONNEMENTS – 2024/2025/2026  |  {d['nom'].upper()}",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols2))
    header_row(ws2, 4, cols2)
    set_widths(ws2, [14,11,26,32,22,10,7,18,20,14,16])

    appros = []
    for year in [2024,2025,2026]:
        n = 25 if year==2024 else (40 if year==2025 else 14)
        for i in range(n):
            dt   = rand_date(year, 5 if year==2026 else 12)
            f_   = random.choice(d["fournisseurs"])
            p    = random.choice(d["produits"])
            qty  = random.randint(50,2000)
            pu   = int(p[4] * random.uniform(0.85,0.97))
            tot  = qty * pu
            sta  = random.choice(STAT_APPRO)
            appros.append((f"APP-{depot_key[:3].upper()}-{year}-{i+1:04d}",
                           dt, f_[1], p[1], p[2], qty, p[3], pu, tot, sta,
                           f"BC-{depot_key[:3].upper()}-{year}-{i+1:04d}"))

    appros.sort(key=lambda x: x[1])
    total_g = sum(a[8] for a in appros)
    for idx, v in enumerate(appros):
        h = {10: GREEN_LIGHT if v[9]=="RÉCEPTIONNÉ" else (YELLOW_L if v[9]=="PARTIEL" else RED_LIGHT)}
        data_row(ws2, 5+idx, v, alt=(idx%2==1), hl=h)
        ws2.row_dimensions[5+idx].height = 18
    total_bar(ws2, 5+len(appros), 8, "TOTAL GÉNÉRAL", 9, total_g)

    path = f"/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Fournisseurs_{depot_key}.xlsx"
    wb.save(path)
    print(f"✓ Rapport_Fournisseurs_{depot_key}.xlsx")


# ════════════════════════════════════════════════════════════════
#  PRODUITS
# ════════════════════════════════════════════════════════════════
def generer_produits(depot_key):
    d   = DEPOTS[depot_key]
    wb  = openpyxl.Workbook(); wb.remove(wb.active)

    stock = {p[0]: random.randint(p[6], p[6]*6) for p in d["produits"]}

    # Catalogue
    ws = wb.create_sheet("Catalogue Produits")
    ws.freeze_panes = "A5"
    cols = ["Code","Désignation","Catégorie","Unité",
            "Prix achat (FCFA)","Prix vente (FCFA)","Marge (FCFA)","Marge (%)",
            "Stock actuel","Stock mini","Fournisseur principal","Statut"]
    banner(ws, f"CATALOGUE DES PRODUITS  |  {d['nom'].upper()}",
           f"AGSV – Référentiel produits  |  Mis à jour le {date.today():%d/%m/%Y}", len(cols))
    header_row(ws, 4, cols)
    set_widths(ws, [10,32,22,7,18,18,14,10,12,10,24,10])

    for idx, p in enumerate(d["produits"]):
        code,nom,cat,unit,p_ach,p_vte,s_min,fourn = p
        marge   = p_vte - p_ach
        marge_p = round(marge/p_ach*100,1)
        stk     = stock[code]
        statut  = "ACTIF" if stk > s_min else "RUPTURE"
        data_row(ws, 5+idx,
                 [code,nom,cat,unit,p_ach,p_vte,marge,f"{marge_p} %",stk,s_min,fourn,statut],
                 alt=(idx%2==1), hl={12: GREEN_LIGHT if statut=="ACTIF" else RED_LIGHT})
        ws.row_dimensions[5+idx].height = 20

    # Stock actuel
    ws2 = wb.create_sheet("État du Stock")
    ws2.freeze_panes = "A5"
    cols2 = ["Code","Désignation","Catégorie","Unité",
             "Stock actuel","Stock mini","Valeur stock (FCFA)","Alerte"]
    banner(ws2, f"ÉTAT DU STOCK  |  {d['nom'].upper()}",
           f"AGSV – Situation au {date.today():%d/%m/%Y}", len(cols2))
    header_row(ws2, 4, cols2)
    set_widths(ws2, [10,32,22,7,14,10,22,16])

    val_total = 0
    for idx, p in enumerate(d["produits"]):
        code,nom,cat,unit,p_ach,p_vte,s_min,fourn = p
        stk   = stock[code]
        val   = stk * p_vte
        val_total += val
        alerte = "⚠ STOCK BAS" if stk<=s_min*1.5 else ("✓ OK" if stk>s_min*3 else "→ MOYEN")
        data_row(ws2, 5+idx, [code,nom,cat,unit,stk,s_min,val,alerte],
                 alt=(idx%2==1),
                 hl={8: RED_LIGHT if "⚠" in alerte else (GREEN_LIGHT if "✓" in alerte else YELLOW_L)})
        ws2.row_dimensions[5+idx].height = 20
    total_bar(ws2, 5+len(d["produits"]), 6, "VALEUR TOTALE", 7, val_total)

    # Mouvements
    ws3 = wb.create_sheet("Mouvements de Stock")
    ws3.freeze_panes = "A5"
    cols3 = ["N° Mouvement","Date","Produit","Catégorie","Type","Quantité","Unité","Motif","Référence"]
    banner(ws3, f"MOUVEMENTS DE STOCK – 2025/2026  |  {d['nom'].upper()}",
           f"AGSV  |  Généré le {date.today():%d/%m/%Y}", len(cols3))
    header_row(ws3, 4, cols3)
    set_widths(ws3, [16,11,32,22,12,10,7,32,14])

    MOTIFS = {
        "ENTRÉE":     ["Approvisionnement fournisseur","Retour client","Transfert reçu"],
        "SORTIE":     ["Vente client","Vente client","Vente client","Don programme"],
        "PERTE":      ["Plants morts en transit","Semences avariées","Détérioration stockage"],
        "INVENTAIRE": ["Inventaire mensuel","Inventaire trimestriel","Ajustement stock"],
    }
    TYPE_C = {"ENTRÉE":GREEN_LIGHT,"SORTIE":WHITE,"PERTE":RED_LIGHT,"INVENTAIRE":YELLOW_L}
    TYPES  = ["ENTRÉE","SORTIE","SORTIE","SORTIE","PERTE","INVENTAIRE"]

    mvts = []
    for year in [2025,2026]:
        for i in range(220 if year==2025 else 85):
            dt  = rand_date(year, 5 if year==2026 else 12)
            p   = random.choice(d["produits"])
            typ = random.choice(TYPES)
            qty = random.randint(5,400)
            mot = random.choice(MOTIFS[typ])
            ref = f"MVT-{depot_key[:3].upper()}-{year}-{i+1:04d}"
            mvts.append((ref,dt,p[1],p[2],typ,qty,p[3],mot,ref.replace("MVT","REF")))

    mvts.sort(key=lambda x: x[1])
    for idx,v in enumerate(mvts):
        data_row(ws3, 5+idx, v, alt=(idx%2==1), hl={5: TYPE_C.get(v[4],WHITE)})
        ws3.row_dimensions[5+idx].height = 18

    # Performance
    ws4 = wb.create_sheet("Performance Produits")
    ws4.freeze_panes = "A5"
    cols4 = ["Code","Désignation","Catégorie",
             "Qté vendue 2025","CA 2025 (FCFA)","Qté vendue 2026 (jan–mai)","CA 2026 (FCFA)",
             "Évolution CA","Rang 2025"]
    banner(ws4, f"PERFORMANCE DES VENTES PAR PRODUIT  |  {d['nom'].upper()}",
           f"AGSV – 2025 vs 2026 (jan–mai)  |  Généré le {date.today():%d/%m/%Y}", len(cols4))
    header_row(ws4, 4, cols4)
    set_widths(ws4, [10,32,22,16,20,22,20,14,10])

    perfs = []
    for p in d["produits"]:
        code,nom,cat,unit,p_ach,p_vte,s_min,fourn = p
        q25  = random.randint(200,3000)
        ca25 = q25 * p_vte
        q26  = random.randint(80,int(q25*0.55))
        ca26 = q26 * p_vte
        evol = round((ca26*12/5 - ca25)/ca25*100,1)
        perfs.append((code,nom,cat,q25,ca25,q26,ca26,evol))

    perfs.sort(key=lambda x: -x[4])
    for idx,(code,nom,cat,q25,ca25,q26,ca26,evol) in enumerate(perfs):
        evol_s = f"▲ +{evol} %" if evol>=0 else f"▼ {evol} %"
        data_row(ws4, 5+idx,
                 [code,nom,cat,q25,ca25,q26,ca26,evol_s,idx+1],
                 alt=(idx%2==1), hl={8: GREEN_LIGHT if evol>=0 else RED_LIGHT})
        ws4.row_dimensions[5+idx].height = 20

    path = f"/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Produits_{depot_key}.xlsx"
    wb.save(path)
    print(f"✓ Rapport_Produits_{depot_key}.xlsx")


# ── Run ───────────────────────────────────────────────────────
for depot in ["Korhogo", "Bamako"]:
    print(f"\n── {depot} ──")
    generer_ventes(depot)
    generer_commandes(depot)
    generer_fournisseurs(depot)
    generer_produits(depot)

print("\nTous les fichiers générés dans /rapports/")