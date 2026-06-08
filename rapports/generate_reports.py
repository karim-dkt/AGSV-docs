import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import random
from datetime import datetime, date, timedelta

random.seed(42)

# ── Palette ────────────────────────────────────────────────────
GREEN_DARK   = "1A6B3A"
GREEN_MED    = "2E7D32"
GREEN_LIGHT  = "E8F5E9"
ORANGE       = "E65100"
ORANGE_LIGHT = "FFF3E0"
GREY_HEAD    = "F5F5F5"
GREY_ROW     = "FAFAFA"
WHITE        = "FFFFFF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="1A1A1A", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def apply_header_row(ws, row, cols):
    for col_idx, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.fill      = _fill(GREEN_DARK)
        c.font      = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = _center()
        c.border    = _border()

def apply_data_row(ws, row, values, alt=False):
    bg = GREY_ROW if alt else WHITE
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col_idx, value=val)
        c.fill   = _fill(bg)
        c.border = _border()
        c.font   = _font()
        if isinstance(val, (int, float)):
            c.alignment = _right()
            if col_idx >= 5:
                c.number_format = '#,##0'
        else:
            c.alignment = _left()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def banner(ws, title, subtitle, merge_to):
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.merge_cells(f"A1:{get_column_letter(merge_to)}1")
    c = ws["A1"]
    c.value     = title
    c.fill      = _fill(GREEN_DARK)
    c.font      = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = _center()
    ws.merge_cells(f"A2:{get_column_letter(merge_to)}2")
    c = ws["A2"]
    c.value     = subtitle
    c.fill      = _fill(GREEN_MED)
    c.font      = Font(italic=True, color="FFFFFF", size=10, name="Calibri")
    c.alignment = _center()

# ── Référentiels ───────────────────────────────────────────────
PRODUITS = [
    ("Riz local 50 kg",        25000, "Alimentation"),
    ("Riz importé 25 kg",      18500, "Alimentation"),
    ("Huile de palme 5 L",      4500, "Alimentation"),
    ("Sucre en poudre 50 kg",  28000, "Alimentation"),
    ("Farine de blé 25 kg",    12000, "Alimentation"),
    ("Lait en poudre 2,5 kg",   8500, "Alimentation"),
    ("Tomate concentrée x24",  15000, "Alimentation"),
    ("Sardines conserve x24",  22000, "Alimentation"),
    ("Café soluble 500 g",      3500, "Boissons"),
    ("Eau minérale x12 1,5 L",  5500, "Boissons"),
    ("Savon de ménage x12",     6000, "Hygiène"),
    ("Détergent 5 kg",          9000, "Hygiène"),
    ("Lessive en poudre 5 kg",  7500, "Hygiène"),
    ("Huile moteur 5 L",       18000, "Quincaillerie"),
    ("Pile alcaline x12",       3000, "Quincaillerie"),
]

CLIENTS = [
    "Restaurant Le Palmier",
    "Supermarché Abidjan Nord",
    "Épicerie Chez Kouamé",
    "Hôtel Ivoire Business",
    "Cantine Scolaire – Lycée Technique",
    "Boutique Aminata Diallo",
    "GIE Femmes de Yopougon",
    "Cafétéria Université FHB",
    "Maquis Le Tropical",
    "Commerce Général Traoré",
    "Superette Bonheur",
    "Alimentaire Ouattara & Fils",
]

VENDEURS = [
    ("Koné Moussa",      "Dépôt Yopougon"),
    ("Diallo Fatoumata", "Dépôt Cocody"),
    ("Bamba Seydou",     "Dépôt Principal"),
    ("Coulibaly Aïcha",  "Dépôt Yopougon"),
]

MODES_PAIEMENT = ["ESPÈCES", "MOBILE MONEY", "VIREMENT", "CHÈQUE"]
STATUTS_VENTE  = ["ENCAISSÉE", "ENCAISSÉE", "ENCAISSÉE", "CRÉDIT"]  # 75 % encaissée
STATUTS_CMD    = ["LIVRÉE", "LIVRÉE", "LIVRÉE", "EN COURS", "ANNULÉE"]

FOURNISSEURS = [
    ("SIFCA",              "Industriel",   "Abidjan – Zone Industrielle", "27 21 24 00 00", "contact@sifca.ci",         "Huile de palme, Savon"),
    ("SANIA Côte d'Ivoire","Industriel",   "San Pedro",                   "27 34 71 12 00", "commercial@sania.ci",       "Riz, Huile végétale"),
    ("Grand Moulin CI",    "Industriel",   "Abidjan – Treichville",       "27 21 75 00 00", "gmc@grandmoulin.ci",        "Farine de blé, Semoule"),
    ("NESTLÉ CI",          "Industriel",   "Abidjan – Plateau",           "27 20 31 00 00", "nestle@nestle.ci",          "Lait en poudre, Café, Eau"),
    ("UNILEVER CI",        "Industriel",   "Abidjan – Zone 4",            "27 21 56 00 00", "ulci@unilever.com",         "Savon, Détergent, Lessive"),
    ("CFCI",               "Industriel",   "Abidjan – Koumassi",          "27 21 35 00 00", "cfci@cfci.ci",              "Conserves, Tomate"),
    ("Diallo Commerce",    "Grossiste",    "Yopougon – Marché",           "07 00 11 22 33", "diallocommerce@gmail.com",  "Épices, Condiments"),
    ("Traoré Distribution","Grossiste",    "Abobo – Centre",              "05 77 88 99 00", "traore.dist@yahoo.fr",      "Boissons, Eau minérale"),
    ("PETROCI",            "Industriel",   "Abidjan – Vridi",             "27 21 08 00 00", "contact@petroci.ci",        "Huile moteur, Lubrifiants"),
    ("ElecPlus CI",        "Grossiste",    "Adjamé – Grand Marché",       "07 44 55 66 77", "elecplus.ci@gmail.com",     "Piles, Éclairage"),
]


def rand_date(year, month_min=1, month_max=12):
    month = random.randint(month_min, month_max)
    if year == 2026 and month > 5:
        month = random.randint(1, 5)
    day_max = 28 if month == 2 else (30 if month in [4,6,9,11] else 31)
    day = random.randint(1, day_max)
    return date(year, month, day)


# ───────────────────────────────────────────────────────────────
#  RAPPORT VENTES
# ───────────────────────────────────────────────────────────────
def build_ventes(wb, year, n_ventes):
    ws_detail = wb.create_sheet(f"Ventes {year}")
    ws_detail.freeze_panes = "A5"

    cols = ["N° Vente", "Date", "Client", "Produit", "Catégorie",
            "Qté", "Prix unit. (FCFA)", "Total ligne (FCFA)",
            "Mode paiement", "Statut", "Vendeur", "Dépôt"]
    banner(ws_detail, f"RAPPORT DÉTAILLÉ DES VENTES – {year}",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols))
    apply_header_row(ws_detail, 4, cols)
    set_col_widths(ws_detail, [12, 12, 30, 28, 16, 6, 18, 18, 16, 12, 22, 18])

    row = 5
    totaux_mensuels = {}
    ventes_data = []

    for i in range(1, n_ventes + 1):
        num = f"VTE-{year}-{i:04d}"
        d   = rand_date(year, 1, 12 if year == 2025 else 5)
        cl  = random.choice(CLIENTS)
        prod, prix_u, cat = random.choice(PRODUITS)
        qty  = random.randint(1, 20)
        total = qty * prix_u
        mode  = random.choice(MODES_PAIEMENT)
        stat  = random.choice(STATUTS_VENTE)
        vend, depot = random.choice(VENDEURS)
        ventes_data.append((num, d, cl, prod, cat, qty, prix_u, total, mode, stat, vend, depot))

    ventes_data.sort(key=lambda x: x[1])

    for idx, vals in enumerate(ventes_data):
        apply_data_row(ws_detail, row, vals, alt=(idx % 2 == 1))
        ws_detail.row_dimensions[row].height = 18
        m = vals[1].month
        totaux_mensuels[m] = totaux_mensuels.get(m, 0) + vals[7]
        row += 1

    # Total général
    total_global = sum(v[7] for v in ventes_data)
    ws_detail.merge_cells(f"A{row}:G{row}")
    c = ws_detail[f"A{row}"]
    c.value = "TOTAL GÉNÉRAL"
    c.font  = _font(bold=True, color="FFFFFF")
    c.fill  = _fill(GREEN_DARK)
    c.alignment = _center()
    c2 = ws_detail[f"H{row}"]
    c2.value = total_global
    c2.font  = _font(bold=True, color="FFFFFF")
    c2.fill  = _fill(GREEN_DARK)
    c2.alignment = _right()
    c2.number_format = '#,##0'

    # ── Résumé mensuel ──
    ws_sum = wb.create_sheet(f"Résumé {year}")
    ws_sum.freeze_panes = "A5"
    MOIS = ["Janvier","Février","Mars","Avril","Mai","Juin",
            "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
    banner(ws_sum, f"RÉSUMÉ MENSUEL DES VENTES – {year}",
           f"AGSV – Consolidation mensuelle  |  Généré le {date.today():%d/%m/%Y}", 4)
    apply_header_row(ws_sum, 4, ["Mois", "Nb transactions", "CA (FCFA)", "% du CA annuel"])
    set_col_widths(ws_sum, [18, 18, 22, 18])

    nb_par_mois = {}
    for v in ventes_data:
        m = v[1].month
        nb_par_mois[m] = nb_par_mois.get(m, 0) + 1

    max_mois = 12 if year == 2025 else 5
    for idx, m in enumerate(range(1, max_mois + 1)):
        ca  = totaux_mensuels.get(m, 0)
        nb  = nb_par_mois.get(m, 0)
        pct = round(ca / total_global * 100, 1) if total_global else 0
        apply_data_row(ws_sum, 4 + idx + 1, [MOIS[m-1], nb, ca, f"{pct} %"], alt=(idx % 2 == 1))
        ws_sum.row_dimensions[4 + idx + 1].height = 18

    total_row = 4 + max_mois + 1
    ws_sum.merge_cells(f"A{total_row}:B{total_row}")
    ws_sum[f"A{total_row}"].value = "TOTAL"
    ws_sum[f"A{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws_sum[f"A{total_row}"].fill  = _fill(GREEN_DARK)
    ws_sum[f"A{total_row}"].alignment = _center()
    ws_sum[f"C{total_row}"].value = total_global
    ws_sum[f"C{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws_sum[f"C{total_row}"].fill  = _fill(GREEN_DARK)
    ws_sum[f"C{total_row}"].alignment = _right()
    ws_sum[f"C{total_row}"].number_format = '#,##0'
    ws_sum[f"D{total_row}"].value = "100 %"
    ws_sum[f"D{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws_sum[f"D{total_row}"].fill  = _fill(GREEN_DARK)
    ws_sum[f"D{total_row}"].alignment = _center()


# ───────────────────────────────────────────────────────────────
#  RAPPORT COMMANDES
# ───────────────────────────────────────────────────────────────
def build_commandes(wb, year, n_cmd):
    ws = wb.create_sheet(f"Commandes {year}")
    ws.freeze_panes = "A5"

    cols = ["N° Commande", "Date commande", "Date livraison prévue",
            "Client", "Produit", "Qté commandée", "Qté livrée",
            "Prix unit. (FCFA)", "Montant (FCFA)", "Statut", "Vendeur", "Dépôt"]
    banner(ws, f"RAPPORT DES COMMANDES – {year}",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols))
    apply_header_row(ws, 4, cols)
    set_col_widths(ws, [14, 14, 20, 30, 28, 14, 12, 18, 18, 12, 22, 18])

    cmds = []
    for i in range(1, n_cmd + 1):
        num     = f"CMD-{year}-{i:04d}"
        d_cmd   = rand_date(year, 1, 12 if year == 2025 else 5)
        d_livr  = d_cmd + timedelta(days=random.randint(2, 14))
        cl      = random.choice(CLIENTS)
        prod, prix_u, _ = random.choice(PRODUITS)
        qty_cmd  = random.randint(5, 50)
        stat     = random.choice(STATUTS_CMD)
        qty_livr = qty_cmd if stat == "LIVRÉE" else (
                   random.randint(0, qty_cmd - 1) if stat == "EN COURS" else 0)
        montant  = qty_livr * prix_u
        vend, depot = random.choice(VENDEURS)
        cmds.append((num, d_cmd, d_livr, cl, prod, qty_cmd, qty_livr,
                     prix_u, montant, stat, vend, depot))

    cmds.sort(key=lambda x: x[1])
    total_global = 0

    for idx, vals in enumerate(cmds):
        apply_data_row(ws, 4 + idx + 1, vals, alt=(idx % 2 == 1))
        ws.row_dimensions[4 + idx + 1].height = 18
        total_global += vals[8]

    total_row = 4 + len(cmds) + 1
    ws.merge_cells(f"A{total_row}:H{total_row}")
    ws[f"A{total_row}"].value = "TOTAL LIVRÉ"
    ws[f"A{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws[f"A{total_row}"].fill  = _fill(GREEN_DARK)
    ws[f"A{total_row}"].alignment = _center()
    ws[f"I{total_row}"].value = total_global
    ws[f"I{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws[f"I{total_row}"].fill  = _fill(GREEN_DARK)
    ws[f"I{total_row}"].alignment = _right()
    ws[f"I{total_row}"].number_format = '#,##0'

    # Résumé par statut
    ws_s = wb.create_sheet(f"Statuts Cmd {year}")
    banner(ws_s, f"COMMANDES PAR STATUT – {year}",
           f"AGSV  |  Généré le {date.today():%d/%m/%Y}", 4)
    apply_header_row(ws_s, 4, ["Statut", "Nb commandes", "Montant total (FCFA)", "% du total"])
    set_col_widths(ws_s, [18, 18, 24, 14])
    stat_map = {}
    for c in cmds:
        s = c[9]
        if s not in stat_map:
            stat_map[s] = [0, 0]
        stat_map[s][0] += 1
        stat_map[s][1] += c[8]
    for idx, (s, (nb, mt)) in enumerate(sorted(stat_map.items())):
        pct = round(mt / total_global * 100, 1) if total_global else 0
        apply_data_row(ws_s, 5 + idx, [s, nb, mt, f"{pct} %"], alt=(idx % 2 == 1))


# ───────────────────────────────────────────────────────────────
#  RAPPORT FOURNISSEURS
# ───────────────────────────────────────────────────────────────
def build_fournisseurs(wb):
    # Fiche fournisseurs
    ws_f = wb.create_sheet("Fournisseurs")
    ws_f.freeze_panes = "A5"
    cols_f = ["Code", "Nom fournisseur", "Type", "Ville / Adresse",
              "Téléphone", "Email", "Produits fournis",
              "Délai livraison (j)", "Conditions paiement", "Statut"]
    banner(ws_f, "RÉPERTOIRE DES FOURNISSEURS",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Mis à jour le {date.today():%d/%m/%Y}",
           len(cols_f))
    apply_header_row(ws_f, 4, cols_f)
    set_col_widths(ws_f, [10, 26, 14, 28, 18, 28, 30, 16, 22, 12])

    CONDITIONS = ["30 jours net", "Paiement à la livraison", "15 jours net",
                  "60 jours net", "Paiement comptant"]
    DELAIS     = [3, 5, 7, 10, 14, 2]
    STATUTS    = ["ACTIF", "ACTIF", "ACTIF", "ACTIF", "INACTIF"]

    for idx, (nom, typ, ville, tel, email, produits) in enumerate(FOURNISSEURS):
        code   = f"FRN-{idx+1:03d}"
        delai  = random.choice(DELAIS)
        cond   = random.choice(CONDITIONS)
        statut = random.choice(STATUTS)
        vals = [code, nom, typ, ville, tel, email, produits, delai, cond, statut]
        apply_data_row(ws_f, 5 + idx, vals, alt=(idx % 2 == 1))
        ws_f.row_dimensions[5 + idx].height = 22

    # Historique approvisionnements
    ws_a = wb.create_sheet("Approvisionnements")
    ws_a.freeze_panes = "A5"
    cols_a = ["N° Appro", "Date", "Fournisseur", "Produit",
              "Qté reçue", "Prix unitaire (FCFA)", "Montant total (FCFA)",
              "Dépôt récepteur", "Statut livraison", "N° Bon commande"]
    banner(ws_a, "HISTORIQUE DES APPROVISIONNEMENTS – 2024 / 2025 / 2026",
           f"AGSV – Application de Gestion de Stock et de Vente  |  Généré le {date.today():%d/%m/%Y}",
           len(cols_a))
    apply_header_row(ws_a, 4, cols_a)
    set_col_widths(ws_a, [14, 12, 26, 28, 10, 20, 22, 22, 18, 16])

    DEPOTS = ["Dépôt Principal", "Dépôt Yopougon", "Dépôt Cocody"]
    STAT_A = ["RÉCEPTIONNÉ", "RÉCEPTIONNÉ", "RÉCEPTIONNÉ", "PARTIEL", "EN ATTENTE"]
    appros = []

    for year in [2024, 2025, 2026]:
        n = 40 if year == 2024 else (60 if year == 2025 else 20)
        for i in range(n):
            d = rand_date(year, 1, 12 if year < 2026 else 5)
            fid = random.randint(0, len(FOURNISSEURS) - 1)
            nom_f = FOURNISSEURS[fid][0]
            prod, prix_u, _ = random.choice(PRODUITS)
            qty   = random.randint(10, 200)
            total = qty * int(prix_u * random.uniform(0.80, 0.92))
            depot = random.choice(DEPOTS)
            stat  = random.choice(STAT_A)
            bon   = f"BC-{year}-{i+1:04d}"
            num   = f"APP-{year}-{i+1:04d}"
            appros.append((num, d, nom_f, prod, qty, prix_u, total, depot, stat, bon))

    appros.sort(key=lambda x: x[1])
    total_global = sum(a[6] for a in appros)

    for idx, vals in enumerate(appros):
        apply_data_row(ws_a, 5 + idx, vals, alt=(idx % 2 == 1))
        ws_a.row_dimensions[5 + idx].height = 18

    total_row = 5 + len(appros)
    ws_a.merge_cells(f"A{total_row}:F{total_row}")
    ws_a[f"A{total_row}"].value = "TOTAL GÉNÉRAL"
    ws_a[f"A{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws_a[f"A{total_row}"].fill  = _fill(GREEN_DARK)
    ws_a[f"A{total_row}"].alignment = _center()
    ws_a[f"G{total_row}"].value = total_global
    ws_a[f"G{total_row}"].font  = _font(bold=True, color="FFFFFF")
    ws_a[f"G{total_row}"].fill  = _fill(GREEN_DARK)
    ws_a[f"G{total_row}"].alignment = _right()
    ws_a[f"G{total_row}"].number_format = '#,##0'


# ───────────────────────────────────────────────────────────────
#  GÉNÉRATION
# ───────────────────────────────────────────────────────────────
# Rapport Ventes
wb_v = openpyxl.Workbook()
wb_v.remove(wb_v.active)
build_ventes(wb_v, 2025, 280)
build_ventes(wb_v, 2026, 130)
wb_v.save("/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Ventes_2025-2026.xlsx")
print("✓ Rapport_Ventes_2025-2026.xlsx")

# Rapport Commandes
wb_c = openpyxl.Workbook()
wb_c.remove(wb_c.active)
build_commandes(wb_c, 2025, 180)
build_commandes(wb_c, 2026, 85)
wb_c.save("/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Commandes_2025-2026.xlsx")
print("✓ Rapport_Commandes_2025-2026.xlsx")

# Rapport Fournisseurs
wb_f = openpyxl.Workbook()
wb_f.remove(wb_f.active)
build_fournisseurs(wb_f)
wb_f.save("/home/karim-diakite/DEV/Projets/AGSV/rapports/Rapport_Fournisseurs.xlsx")
print("✓ Rapport_Fournisseurs.xlsx")

print("\nDone — fichiers dans /rapports/")